from django.contrib import admin
from django.urls import path
from django.shortcuts import render
from django import forms
from django.utils import timezone
from django.db.models import Sum, Count, F, Q
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.http import HttpResponse
import csv
from decimal import Decimal
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.conf import settings
import os
import calendar

from kliknijidlo.pdf_utils import czech_pdf_styles, decimal_cs, money_cs, safe_table

User = get_user_model()

try:
    import openpyxl
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from objednavky.models import Order, OrderItem
from jidelnicek.models import DruhJidla
from .models import ReportDummy
from datetime import timedelta
from collections import defaultdict


def fmt_max2(value):
    return decimal_cs(value or 0, places=2, trim=True)




PERIOD_CHOICES = [
    ('today', 'Dnes'),
    ('yesterday', 'Včera'),
    ('week', 'Minulý týden'),
    ('month', 'Minulý měsíc'),
    ('current_month', 'Aktuální měsíc'),
    ('year', 'Aktuální rok'),
    ('custom', 'Vlastní období'),
]

GROUPING_CHOICES = [
    ('day', 'Po dnech'),
    ('total', 'Celkem'),
]

MONTH_CHOICES = [
    (1, 'Leden'),
    (2, 'Únor'),
    (3, 'Březen'),
    (4, 'Duben'),
    (5, 'Květen'),
    (6, 'Červen'),
    (7, 'Červenec'),
    (8, 'Srpen'),
    (9, 'Září'),
    (10, 'Říjen'),
    (11, 'Listopad'),
    (12, 'Prosinec'),
]


def get_previous_month_year(base_date=None):
    current = base_date or timezone.localdate()
    if current.month == 1:
        return 12, current.year - 1
    return current.month - 1, current.year


def food_type_short_label(name):
    value = (name or "").strip()
    digits = "".join(char for char in value if char.isdigit())
    words = [word for word in value.replace(".", " ").split() if not word.isdigit()]
    if not words:
        return value[:3].upper()
    if len(words) == 1:
        token = words[0][:2]
    else:
        token = "".join(word[:1] for word in words[:2])
    return f"{digits}{token.upper()}"[:4]


def report_period_label(form):
    if not form.is_valid():
        return "Bez určeného období"

    period = form.cleaned_data.get('period')
    period_map = dict(PERIOD_CHOICES)
    if period == 'custom':
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        if date_from and date_to:
            return f"{date_from.strftime('%d.%m.%Y')} až {date_to.strftime('%d.%m.%Y')}"
        return "Vlastní období"

    return period_map.get(period, "Vybrané období")


def pdf_pairs_to_rows(pairs, columns=2):
    rows = []
    chunk_size = max(1, columns)
    for index in range(0, len(pairs), chunk_size):
        chunk = pairs[index:index + chunk_size]
        row = []
        for label, value in chunk:
            row.extend([label, value])
        while len(row) < chunk_size * 2:
            row.extend(["", ""])
        rows.append(row)
    return rows


def pdf_summary_pairs(report_type, totals):
    if report_type == 'attendance_matrix':
        return [
            ("Strávníci", str(totals.get('users_count', 0))),
            ("Porce", str(totals.get('total_portions', 0))),
            ("Dotované porce", str(totals.get('subsidized_portions', 0))),
            ("Dotace", money_cs(totals.get('subsidy_total', 0))),
            ("Cena po dotaci", money_cs(totals.get('paid_total', 0))),
        ]
    if report_type == 'subsidy_finance':
        return [
            ("Řádků", str(totals.get('rows_count', 0))),
            ("Porce", str(totals.get('total_portions', 0))),
            ("Dotované porce", str(totals.get('subsidized_portions', 0))),
            ("Nedotovaná cena", money_cs(totals.get('full_price_total', 0))),
            ("Dotace", money_cs(totals.get('subsidy_total', 0))),
            ("Cena po dotaci", money_cs(totals.get('paid_total', 0))),
        ]
    if report_type == 'food_types':
        return [
            ("Řádků", str(totals.get('rows_count', 0))),
            ("Porce", str(totals.get('total_portions', 0))),
            ("Plná cena", money_cs(totals.get('unclaimed_total', 0))),
            ("Dotace", money_cs(totals.get('dotace', 0))),
            ("K platbě", money_cs(totals.get('final_price', 0))),
        ]
    if report_type == 'items':
        return [
            ("Řádků", str(totals.get('rows_count', 0))),
            ("Položek", str(totals.get('total_items', 0))),
            ("Plná cena", money_cs(totals.get('unclaimed_total', 0))),
            ("Dotace", money_cs(totals.get('dotace', 0))),
            ("K platbě", money_cs(totals.get('final_price', 0))),
        ]
    return [
        ("Řádků", str(totals.get('rows_count', 0))),
        ("Objednávek", str(totals.get('orders_count', totals.get('rows_count', 0)))),
        ("Plná cena", money_cs(totals.get('unclaimed_total', 0))),
        ("Dotace", money_cs(totals.get('dotace', 0))),
        ("K platbě", money_cs(totals.get('final_price', 0))),
    ]



class ReportForm(forms.Form):
    previous_month, previous_year = get_previous_month_year()
    month = forms.TypedChoiceField(
        choices=MONTH_CHOICES,
        coerce=int,
        required=False,
        initial=previous_month,
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    year = forms.IntegerField(
        required=False,
        initial=previous_year,
        min_value=2020,
        max_value=2100,
        widget=forms.NumberInput(attrs={'class': 'form-control', 'step': 1})
    )
    period = forms.ChoiceField(
        choices=PERIOD_CHOICES,
        initial='today',
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    grouping = forms.ChoiceField(
        choices=GROUPING_CHOICES,
        initial='day',
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    group = forms.ModelChoiceField(
        queryset=Group.objects.all(),
        required=False,
        empty_label='Všechny skupiny',
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    customer = forms.ModelChoiceField(
        queryset=get_user_model().objects.all(),
        required=False,
        empty_label='Všichni zákazníci (vyhledávání)',
        widget=forms.Select(attrs={'class': 'form-control select2-search'})
    )
    date_from = forms.DateField(
        required=False,
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
    )
    date_to = forms.DateField(
        required=False,
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
    )
    # NOVÉ: Multi-select pro druhy jídel
    food_types = forms.ModelMultipleChoiceField(
        queryset=DruhJidla.objects.all(),
        required=False,
        widget=forms.SelectMultiple(attrs={'class': 'form-control', 'size': '5'}),
        label='Druhy jídel'
    )
    # NOVÉ: Fulltext search
    search = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Hledat...'}),
        label='Vyhledávání'
    )



@admin.register(ReportDummy)
class ReportAdmin(admin.ModelAdmin):
    change_list_template = 'admin/reporty/dashboard.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [path('', self.admin_site.admin_view(self.dashboard_view), name='report_dashboard')]
        return custom_urls + urls

    def resolve_period_range(self, form):
        today = timezone.now().date()
        if not form.is_valid():
            return None, None

        period = form.cleaned_data['period']
        if period == 'today':
            return today, today
        if period == 'yesterday':
            previous_day = today - timedelta(days=1)
            return previous_day, previous_day
        if period == 'week':
            return today - timedelta(days=7), today
        if period == 'month':
            previous_month_end = today.replace(day=1) - timedelta(days=1)
            return previous_month_end.replace(day=1), previous_month_end
        if period == 'current_month':
            return today.replace(day=1), today
        if period == 'year':
            return today.replace(month=1, day=1), today
        return form.cleaned_data.get('date_from'), form.cleaned_data.get('date_to')


    def get_report_calculations(self, form):
        """Pomocná metoda pro výpočet dat reportu sdílená mezi view a exportem"""
        today = timezone.now().date()
        report_data = []
        totals = {'unclaimed_total': 0, 'dotace': 0, 'final_price': 0}
        grouping = 'day'

        if not form.is_valid():
            return report_data, totals, grouping

        period = form.cleaned_data['period']
        grouping = form.cleaned_data.get('grouping', 'day')

        # výpočet období
        if period == 'today':
            date_from = date_to = today
        elif period == 'yesterday':
            date_from = date_to = today - timedelta(days=1)
        elif period == 'week':
            date_from = date_to = today - timedelta(days=7)
        elif period == 'month':
            date_from = date_to = today.replace(day=1) - timedelta(days=1)
        elif period == 'current_month':
            date_from = today.replace(day=1)
            date_to = today
        elif period == 'year':
            date_from = today.replace(month=1, day=1)
            date_to = today
        else:  # custom
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')

        # základní queryset
        queryset = Order.objects.filter(
            status__in=['vydano', 'nevyzvednuto']
        ).select_related('user').prefetch_related('items').order_by('datum_vydeje', 'user__first_name')

        if date_from and date_to:
            queryset = queryset.filter(datum_vydeje__range=[date_from, date_to])

        if form.cleaned_data['group']:
            queryset = queryset.filter(user__groups=form.cleaned_data['group'])

        if form.cleaned_data['customer']:
            queryset = queryset.filter(user=form.cleaned_data['customer'])

        for order in queryset:
            if order.status == 'vydano':
                items = order.items.filter(vydano=True)
            elif order.status == 'nevyzvednuto':
                items = order.items.all()
            else:
                continue

            if not items.exists():
                continue

            unclaimed_total = sum(
                item.quantity * getattr(item.menu_item.jidlo, 'cena', 0) for item in items
            )
            dotace = sum(
                (getattr(item.menu_item.jidlo, 'cena', 0) - item.cena) * item.quantity
                for item in items
            )
            final_price = sum(item.quantity * item.cena for item in items)

            row = {
                'user': order.user.get_full_name(),
                'osobni_cislo': order.user.osobni_cislo or '',
                'identifikacni_medium': order.user.identifikacni_medium or '',
                'date': order.datum_vydeje,
                'status': order.status,
                'unclaimed_total': round(unclaimed_total, 2),
                'dotace': round(dotace, 2),
                'final_price': round(final_price, 2),
            }
            report_data.append(row)
            totals['unclaimed_total'] += unclaimed_total
            totals['dotace'] += dotace
            totals['final_price'] += final_price


        if grouping == 'total':
            grouped = defaultdict(
                lambda: {'unclaimed_total': 0, 'dotace': 0, 'final_price': 0, 'count': 0, 'osobni_cislo': '', 'identifikacni_medium': ''}
            )
            
            for row in report_data:
                key = row['user']
                grouped[key]['unclaimed_total'] += row['unclaimed_total']
                grouped[key]['dotace'] += row['dotace']
                grouped[key]['final_price'] += row['final_price']
                grouped[key]['count'] += 1
                if not grouped[key]['osobni_cislo']:
                    grouped[key]['osobni_cislo'] = row['osobni_cislo']
                if not grouped[key]['identifikacni_medium']:
                    grouped[key]['identifikacni_medium'] = row['identifikacni_medium']
            
            report_data = [{
                'user': user,
                'osobni_cislo': data['osobni_cislo'],
                'identifikacni_medium': data['identifikacni_medium'],
                'unclaimed_total': round(data['unclaimed_total'], 2),
                'dotace': round(data['dotace'], 2),
                'final_price': round(data['final_price'], 2),
                'count': data['count'],
                'grouped': True
            } for user, data in sorted(grouped.items(), key=lambda x: x[1]['final_price'], reverse=True)]

        # NOVÉ: Fulltext search
        search_query = form.cleaned_data.get('search', '').strip()
        if search_query:
            filtered_data = []
            for row in report_data:
                searchable_text = ' '.join([
                    str(row.get('user', '')),
                    str(row.get('osobni_cislo', '')),
                    str(row.get('identifikacni_medium', '')),
                    str(row.get('date', '')),
                    str(row.get('status', '')),
                    str(row.get('unclaimed_total', '')),
                    str(row.get('dotace', '')),
                    str(row.get('final_price', '')),
                ]).lower()
                
                if search_query.lower() in searchable_text:
                    filtered_data.append(row)
            
            report_data = filtered_data

        totals['rows_count'] = len(report_data)
        totals['orders_count'] = sum(row.get('count', 1) for row in report_data) if grouping == 'total' else len(report_data)
        totals['rows_count'] = len(report_data)
        totals['rows_count'] = len(report_data)
        return report_data, totals, grouping

    def build_pdf_document_meta(self, form, report_type, grouping, totals):
        current_time = timezone.localtime(timezone.now())
        selected_group = form.cleaned_data.get('group') if form.is_valid() else None
        selected_customer = form.cleaned_data.get('customer') if form.is_valid() else None
        selected_food_types = list(form.cleaned_data.get('food_types') or []) if form.is_valid() else []
        search_query = form.cleaned_data.get('search', '').strip() if form.is_valid() else ''

        if report_type == 'attendance_matrix' and form.is_valid():
            period_label = f"{dict(MONTH_CHOICES).get(form.cleaned_data.get('month'), form.cleaned_data.get('month'))} {form.cleaned_data.get('year')}"
        else:
            period_label = report_period_label(form)

        grouping_label = 'Měsíční přehled' if report_type == 'attendance_matrix' else ('Po dnech' if grouping == 'day' else 'Celkem')
        return {
            'generated_at': current_time.strftime('%d.%m.%Y %H:%M:%S'),
            'period_label': period_label,
            'grouping_label': grouping_label,
            'group_label': selected_group.name if selected_group else 'Všechny skupiny',
            'customer_label': (selected_customer.get_full_name() or selected_customer.username) if selected_customer else 'Všichni zákazníci',
            'food_types_label': ', '.join(food_type.nazev for food_type in selected_food_types) if selected_food_types else 'Všechny druhy jídel',
            'search_label': search_query or 'Bez fulltextu',
        }

    def build_report_pdf(self, report_type, grouping, form, report_data, totals, filename_base):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'

        styles, font_name = czech_pdf_styles()
        color_green = colors.HexColor('#28a745')
        color_orange = colors.HexColor('#fd7e14')
        color_border = colors.HexColor('#d9e5d2')
        color_soft = colors.HexColor('#f7fbf5')
        color_soft_alt = colors.HexColor('#fcfdfc')
        color_total = colors.HexColor('#fff3e0')

        page_size = landscape(A3) if report_type == 'attendance_matrix' else landscape(A4)
        doc = SimpleDocTemplate(
            response,
            pagesize=page_size,
            rightMargin=0.55 * cm,
            leftMargin=0.55 * cm,
            topMargin=0.9 * cm,
            bottomMargin=0.9 * cm,
        )

        title_map = {
            'amounts': 'Přehled objednávek',
            'items': 'Přehled položek objednávek',
            'food_types': 'Přehled podle druhů jídel',
            'attendance_matrix': 'Čárkovnice a finanční souhrn',
            'subsidy_finance': 'Finanční report dotací',
        }
        subtitle_map = {
            'amounts': 'Souhrnný dokument pro kontrolu objednávek, dotací a plateb.',
            'items': 'Detailní rozpad vydaných položek včetně cen a dotací.',
            'food_types': 'Přehled čerpání po druzích jídel a konkrétních pokrmech.',
            'attendance_matrix': 'Měsíční čárkovnice skutečně vydané stravy doplněná o finanční souhrn.',
            'subsidy_finance': 'Podklad pro účetní a auditní kontrolu dotovaných porcí a cen.',
        }

        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=17,
            leading=20,
            textColor=color_orange,
            spaceAfter=6,
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=9,
            leading=11,
            textColor=colors.HexColor('#516151'),
            spaceAfter=10,
        )
        note_style = ParagraphStyle(
            'ReportNote',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#667266'),
        )

        story = []
        story.append(Paragraph(title_map.get(report_type, 'Report'), title_style))
        story.append(Paragraph(subtitle_map.get(report_type, ''), subtitle_style))

        meta = self.build_pdf_document_meta(form, report_type, grouping, totals)
        filter_rows = pdf_pairs_to_rows([
            ('Období', meta['period_label']),
            ('Seskupení', meta['grouping_label']),
            ('Skupina', meta['group_label']),
            ('Zákazník', meta['customer_label']),
            ('Druhy jídel', meta['food_types_label']),
            ('Hledání', meta['search_label']),
        ], columns=2)
        filter_widths = [2.1 * cm, 9.1 * cm, 2.3 * cm, 9.1 * cm] if report_type == 'attendance_matrix' else [2.0 * cm, 7.1 * cm, 2.1 * cm, 7.1 * cm]
        story.append(
            safe_table(
                filter_rows,
                filter_widths,
                font_name=font_name,
                font_size=8,
                header=False,
                repeat_rows=0,
                style_commands=[
                    ('BACKGROUND', (0, 0), (-1, -1), color_soft),
                    ('GRID', (0, 0), (-1, -1), 0.35, color_border),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ],
            )
        )
        story.append(Spacer(1, 0.22 * cm))

        summary_rows = pdf_pairs_to_rows(pdf_summary_pairs(report_type, totals), columns=3 if report_type != 'attendance_matrix' else 2)
        summary_widths = [3.0 * cm, 5.0 * cm, 3.0 * cm, 5.0 * cm, 3.0 * cm, 5.0 * cm]
        if report_type == 'attendance_matrix':
            summary_widths = [3.0 * cm, 7.4 * cm, 3.0 * cm, 7.4 * cm]
        story.append(
            safe_table(
                summary_rows,
                summary_widths,
                font_name=font_name,
                font_size=8,
                header=False,
                repeat_rows=0,
                style_commands=[
                    ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.35, color_border),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ],
            )
        )
        story.append(Spacer(1, 0.28 * cm))

        if report_type == 'attendance_matrix':
            matrix_headers = ['Strávník', 'Druh jídla'] + [str(day) for day in totals['days']] + ['Porcí', 'Dot. porcí', 'Nedot.', 'Dotace', 'Po dotaci']
            matrix_rows = [matrix_headers]
            for row in report_data:
                first_row_for_user = True
                for legend_item in totals.get('legend', []):
                    food_type_name = legend_item['name']
                    label = legend_item['label']
                    finance = row['financial_by_type'].get(food_type_name)
                    if not finance or not finance['count']:
                        continue
                    matrix_rows.append(
                        [
                            (
                                f"{row['user']}\nOsobní číslo: {row['osobni_cislo'] or '-'}\nID médium: {row['identifikacni_medium'] or '-'}"
                            ) if first_row_for_user else '',
                            f"{label} • {food_type_name}",
                        ] + [
                            ('✓' * row['days'][day].count(label)) or ''
                            for day in totals['days']
                        ] + [
                            str(finance['count']),
                            str(finance['subsidized_count']),
                            money_cs(finance['full_price_total']),
                            money_cs(finance['subsidy_total']),
                            money_cs(finance['paid_total']),
                        ]
                    )
                    first_row_for_user = False
            matrix_rows.append(
                ['CELKEM', 'Porce'] + [str(totals['day_totals'][day]['portions']) for day in totals['days']] + [
                    str(totals['total_portions']),
                    str(totals['subsidized_portions']),
                    money_cs(totals['full_price_total']),
                    money_cs(totals['subsidy_total']),
                    money_cs(totals['paid_total']),
                ]
            )
            matrix_rows.append(
                ['', 'Strávníci'] + [str(totals['day_totals'][day]['users']) for day in totals['days']] + [str(totals['users_count']), '', '', '', '']
            )

            matrix_col_widths = [4.4 * cm, 2.8 * cm] + [0.52 * cm for _ in totals['days']] + [1.05 * cm, 1.35 * cm, 1.9 * cm, 1.7 * cm, 1.9 * cm]
            attendance_style = [
                ('BACKGROUND', (0, 0), (-1, 0), color_green),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.3, color_border),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (-5, 1), (-1, -1), 'RIGHT'),
                ('BACKGROUND', (0, -2), (-1, -1), color_total),
                ('LINEABOVE', (0, -2), (-1, -2), 0.8, color_orange),
            ]
            for row_index in range(1, len(matrix_rows) - 2):
                attendance_style.append(('BACKGROUND', (0, row_index), (-1, row_index), color_soft if row_index % 2 else color_soft_alt))
            story.append(
                safe_table(
                    matrix_rows,
                    matrix_col_widths,
                    font_name=font_name,
                    font_size=6,
                    repeat_rows=1,
                    style_commands=attendance_style,
                )
            )
            story.append(Spacer(1, 0.24 * cm))

            if totals.get('financial_rows'):
                finance_rows = [[
                    'Druh jídla', 'Zkratka', 'Porcí', 'Dotovaných porcí', 'Nedotovaná cena', 'Dotace', 'Cena po dotaci',
                ]]
                for finance_row in totals['financial_rows']:
                    finance_rows.append([
                        finance_row['food_type'],
                        finance_row['label'],
                        str(finance_row['count']),
                        str(finance_row['subsidized_count']),
                        money_cs(finance_row['full_price_total']),
                        money_cs(finance_row['subsidy_total']),
                        money_cs(finance_row['paid_total']),
                    ])
                finance_rows.append([
                    'CELKEM', '', str(totals['total_portions']), str(totals['subsidized_portions']),
                    money_cs(totals['full_price_total']), money_cs(totals['subsidy_total']), money_cs(totals['paid_total']),
                ])
                finance_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), color_green),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.35, color_border),
                    ('BACKGROUND', (0, -1), (-1, -1), color_total),
                    ('LINEABOVE', (0, -1), (-1, -1), 0.8, color_orange),
                    ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                    ('ALIGN', (-3, 1), (-1, -1), 'RIGHT'),
                ]
                for row_index in range(1, len(finance_rows) - 1):
                    finance_style.append(('BACKGROUND', (0, row_index), (-1, row_index), color_soft if row_index % 2 else color_soft_alt))
                story.append(Paragraph("Finanční souhrn podle druhů jídel", styles['Heading2']))
                story.append(Spacer(1, 0.12 * cm))
                story.append(
                    safe_table(
                        finance_rows,
                        [4.2 * cm, 1.7 * cm, 1.5 * cm, 2.2 * cm, 2.9 * cm, 2.6 * cm, 2.9 * cm],
                        font_name=font_name,
                        font_size=7,
                        repeat_rows=1,
                        style_commands=finance_style,
                    )
                )
            if totals.get('legend'):
                story.append(Spacer(1, 0.18 * cm))
                legend_text = " • ".join(f"{item['label']} = {item['name']}" for item in totals['legend'])
                story.append(Paragraph(f"<b>Legenda:</b> {legend_text}", note_style))
        else:
            if report_type == 'subsidy_finance':
                if grouping == 'day':
                    headers = ['Datum', 'Zákazník', 'Druh', 'Jídlo', 'Porcí', 'Dot. porcí', 'Nedot.', 'Dotace', 'Po dotaci']
                    col_widths = [2.2 * cm, 3.8 * cm, 2.3 * cm, 5.8 * cm, 1.2 * cm, 1.6 * cm, 2.4 * cm, 2.4 * cm, 2.4 * cm]
                else:
                    headers = ['Zákazník', 'Druhy jídel', 'Porcí', 'Dot. porcí', 'Nedot.', 'Dotace', 'Po dotaci']
                    col_widths = [4.4 * cm, 6.5 * cm, 1.6 * cm, 1.9 * cm, 2.7 * cm, 2.7 * cm, 2.7 * cm]
            elif report_type == 'food_types':
                if grouping == 'day':
                    headers = ['Datum', 'Zákazník', 'Druh jídla', 'Název jídla', 'Ks', 'Plná', 'Dotace', 'K platbě']
                    col_widths = [2.2 * cm, 3.8 * cm, 3.0 * cm, 6.8 * cm, 1.1 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm]
                else:
                    headers = ['Zákazník', 'Druh jídla', 'Název jídla', 'Ks', 'Plná', 'Dotace', 'K platbě']
                    col_widths = [4.0 * cm, 3.0 * cm, 8.0 * cm, 1.2 * cm, 2.4 * cm, 2.4 * cm, 2.4 * cm]
            else:
                if grouping == 'day':
                    headers = ['Zákazník', 'Osobní č.', 'ID médium', 'Datum', 'Stav', 'Plná', 'Dotace', 'K platbě']
                    col_widths = [4.0 * cm, 2.3 * cm, 2.7 * cm, 2.4 * cm, 2.2 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm]
                else:
                    headers = ['Zákazník', 'Osobní č.', 'ID médium', 'Počet', 'Plná', 'Dotace', 'K platbě']
                    col_widths = [5.0 * cm, 2.8 * cm, 3.0 * cm, 1.7 * cm, 2.7 * cm, 2.7 * cm, 2.7 * cm]

            table_rows = [headers]
            for row in report_data:
                if report_type == 'subsidy_finance':
                    line = [row['date'].strftime('%d.%m.%Y'), row['user']] if grouping == 'day' else [row['user']]
                    if grouping == 'day':
                        line += [row['food_type'], row['food_name'], str(row['quantity']), str(row['subsidized_portions'])]
                    else:
                        line += [row.get('food_types', ''), str(row['total_portions']), str(row['subsidized_portions'])]
                    line += [money_cs(row['full_price_total']), money_cs(row['subsidy_total']), money_cs(row['paid_total'])]
                elif report_type == 'food_types':
                    line = [row['date'].strftime('%d.%m.%Y')] if grouping == 'day' else []
                    line += [row['user'], row['food_type'], row['food_name'], str(row['quantity']), money_cs(row['unclaimed_total']), money_cs(row['dotace']), money_cs(row['final_price'])]
                else:
                    line = [row['user'], row['osobni_cislo'], row['identifikacni_medium']]
                    if grouping == 'day':
                        line += [row['date'].strftime('%d.%m.%Y'), row['status']]
                    else:
                        line += [f"{row.get('count', row.get('items_count', 0))} ks"]
                    line += [money_cs(row['unclaimed_total']), money_cs(row['dotace']), money_cs(row['final_price'])]
                table_rows.append(line)

            if report_type == 'subsidy_finance':
                footer = ['CELKEM'] + ([''] * (5 if grouping == 'day' else 3)) + [money_cs(totals['full_price_total']), money_cs(totals['subsidy_total']), money_cs(totals['paid_total'])]
            elif report_type == 'food_types':
                footer = ['CELKEM'] + ([''] if grouping == 'day' else []) + ['', '', str(totals['total_portions']), money_cs(totals['unclaimed_total']), money_cs(totals['dotace']), money_cs(totals['final_price'])]
            else:
                footer = ['CELKEM'] + (['', '', '', ''] if grouping == 'day' else ['', '', '']) + [money_cs(totals['unclaimed_total']), money_cs(totals['dotace']), money_cs(totals['final_price'])]
            table_rows.append(footer)

            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), color_green),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.35, color_border),
                ('BACKGROUND', (0, -1), (-1, -1), color_total),
                ('LINEABOVE', (0, -1), (-1, -1), 0.8, color_orange),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (-3, 1), (-1, -1), 'RIGHT'),
            ]
            for row_index in range(1, len(table_rows) - 1):
                table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), color_soft if row_index % 2 else color_soft_alt))
            story.append(
                safe_table(
                    table_rows,
                    col_widths,
                    font_name=font_name,
                    font_size=7,
                    repeat_rows=1,
                    style_commands=table_style,
                )
            )

        story.append(Spacer(1, 0.22 * cm))
        story.append(Paragraph(f"<i>Vygenerováno: {meta['generated_at']} (Praha)</i>", note_style))
        doc.build(story)
        return response


    def get_order_items_report(self, form):
        """Report pro položky objednávek - počet a názvy jídel"""
        today = timezone.now().date()
        report_data = []
        totals = {'total_items': 0, 'unclaimed_total': 0, 'dotace': 0, 'final_price': 0}
        grouping = 'day'
        
        if not form.is_valid():
            return report_data, totals, grouping
        
        period = form.cleaned_data['period']
        grouping = form.cleaned_data.get('grouping', 'day')
        
        # výpočet období
        if period == 'today':
            date_from = date_to = today
        elif period == 'yesterday':
            date_from = date_to = today - timedelta(days=1)
        elif period == 'week':
            date_from = date_to = today - timedelta(days=7)
        elif period == 'month':
            date_from = date_to = today.replace(day=1) - timedelta(days=1)
        elif period == 'current_month':
            date_from = today.replace(day=1)
            date_to = today
        elif period == 'year':
            date_from = today.replace(month=1, day=1)
            date_to = today
        else:  # custom
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
        
        # Queryset pro OrderItems
        queryset = OrderItem.objects.filter(
            order__status__in=['vydano', 'nevyzvednuto'],
            vydano=True
        ).select_related('order__user', 'menu_item__jidlo').order_by('order__datum_vydeje', 'order__user__first_name')
        
        if date_from and date_to:
            queryset = queryset.filter(order__datum_vydeje__range=[date_from, date_to])
        
        if form.cleaned_data['group']:
            queryset = queryset.filter(order__user__groups=form.cleaned_data['group'])
        
        if form.cleaned_data['customer']:
            queryset = queryset.filter(order__user=form.cleaned_data['customer'])
        
        for item in queryset:
            jidlo_nazev = getattr(item.menu_item.jidlo, 'nazev', 'N/A')
            jidlo_cena = getattr(item.menu_item.jidlo, 'cena', 0)
            
            unclaimed_total = item.quantity * jidlo_cena
            dotace = (jidlo_cena - item.cena) * item.quantity
            final_price = item.quantity * item.cena
            
            row = {
                'user': item.order.user.get_full_name(),
                'osobni_cislo': item.order.user.osobni_cislo or '',
                'identifikacni_medium': item.order.user.identifikacni_medium or '',
                'date': item.order.datum_vydeje,
                'status': item.order.status,
                'jidlo_nazev': jidlo_nazev,
                'quantity': item.quantity,
                'unclaimed_total': round(unclaimed_total, 2),
                'dotace': round(dotace, 2),
                'final_price': round(final_price, 2),
            }
            report_data.append(row)
            
            totals['total_items'] += item.quantity
            totals['unclaimed_total'] += unclaimed_total
            totals['dotace'] += dotace
            totals['final_price'] += final_price
        
        if grouping == 'total':
            grouped = defaultdict(
                lambda: {
                    'unclaimed_total': 0, 
                    'dotace': 0, 
                    'final_price': 0, 
                    'items_count': 0,
                    'items_list': [],
                    'osobni_cislo': '', 
                    'identifikacni_medium': ''
                }
            )
            
            for row in report_data:
                key = row['user']
                grouped[key]['unclaimed_total'] += row['unclaimed_total']
                grouped[key]['dotace'] += row['dotace']
                grouped[key]['final_price'] += row['final_price']
                grouped[key]['items_count'] += row['quantity']
                grouped[key]['items_list'].append(f"{row['jidlo_nazev']} ({row['quantity']}x)")
                
                if not grouped[key]['osobni_cislo']:
                    grouped[key]['osobni_cislo'] = row['osobni_cislo']
                if not grouped[key]['identifikacni_medium']:
                    grouped[key]['identifikacni_medium'] = row['identifikacni_medium']
            
            report_data = [{
                'user': user,
                'osobni_cislo': data['osobni_cislo'],
                'identifikacni_medium': data['identifikacni_medium'],
                'items_count': data['items_count'],
                'items_names': ', '.join(data['items_list']),
                'unclaimed_total': round(data['unclaimed_total'], 2),
                'dotace': round(data['dotace'], 2),
                'final_price': round(data['final_price'], 2),
                'grouped': True
            } for user, data in sorted(grouped.items(), key=lambda x: x[1]['final_price'], reverse=True)]

        # NOVÉ: Fulltext search
        search_query = form.cleaned_data.get('search', '').strip()
        if search_query:
            filtered_data = []
            for row in report_data:
                searchable_text = ' '.join([
                    str(row.get('user', '')),
                    str(row.get('osobni_cislo', '')),
                    str(row.get('identifikacni_medium', '')),
                    str(row.get('date', '')),
                    str(row.get('status', '')),
                    str(row.get('jidlo_nazev', '')),
                    str(row.get('quantity', '')),
                    str(row.get('items_names', '')),
                    str(row.get('unclaimed_total', '')),
                    str(row.get('dotace', '')),
                    str(row.get('final_price', '')),
                ]).lower()
                
                if search_query.lower() in searchable_text:
                    filtered_data.append(row)
            
            report_data = filtered_data
        
        totals['rows_count'] = len(report_data)
        return report_data, totals, grouping


    def get_food_types_report(self, form):
        """NOVÝ: Report pro druhy jídel - detailní výpis s jmény zákazníků a názvy jídel"""
        today = timezone.now().date()
        report_data = []
        totals = {'total_portions': 0, 'unclaimed_total': 0, 'dotace': 0, 'final_price': 0}
        grouping = 'day'
        
        if not form.is_valid():
            return report_data, totals, grouping
        
        period = form.cleaned_data['period']
        grouping = form.cleaned_data.get('grouping', 'day')
        
        # výpočet období
        if period == 'today':
            date_from = date_to = today
        elif period == 'yesterday':
            date_from = date_to = today - timedelta(days=1)
        elif period == 'week':
            date_from = date_to = today - timedelta(days=7)
        elif period == 'month':
            date_from = date_to = today.replace(day=1) - timedelta(days=1)
        elif period == 'current_month':
            date_from = today.replace(day=1)
            date_to = today
        elif period == 'year':
            date_from = today.replace(month=1, day=1)
            date_to = today
        else:  # custom
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
        
        # Základní queryset - detailní položky
        queryset = OrderItem.objects.filter(
            order__status__in=['vydano', 'nevyzvednuto'],
            
        ).select_related('order__user', 'menu_item__jidlo', 'menu_item__druh_jidla').order_by(
            'order__datum_vydeje', 'order__user__first_name', 'menu_item__druh_jidla__nazev'
        )
        
        if date_from and date_to:
            queryset = queryset.filter(order__datum_vydeje__range=[date_from, date_to])
        
        if form.cleaned_data['group']:
            queryset = queryset.filter(order__user__groups=form.cleaned_data['group'])
        
        if form.cleaned_data['customer']:
            queryset = queryset.filter(order__user=form.cleaned_data['customer'])
        
        # Filtr podle druhů jídel
        if form.cleaned_data.get('food_types'):
            queryset = queryset.filter(menu_item__druh_jidla__in=form.cleaned_data['food_types'])
        
        # Procházej jednotlivé položky
        if grouping == 'day':
            for item in queryset:
                jidlo_nazev = getattr(item.menu_item.jidlo, 'nazev', 'N/A')
                jidlo_cena = getattr(item.menu_item.jidlo, 'cena', 0)
                druh_jidla = getattr(item.menu_item.druh_jidla, 'nazev', 'Neurčeno')
                
                unclaimed_total = item.quantity * jidlo_cena
                dotace = (jidlo_cena - item.cena) * item.quantity
                final_price = item.quantity * item.cena
                
                row = {
                    'date': item.order.datum_vydeje,
                    'user': item.order.user.get_full_name(),
                    'food_type': druh_jidla,
                    'food_name': jidlo_nazev,
                    'quantity': item.quantity,
                    'unclaimed_total': round(unclaimed_total, 2),
                    'dotace': round(dotace, 2),
                    'final_price': round(final_price, 2),
                }
                report_data.append(row)
                
                totals['total_portions'] += item.quantity
                totals['unclaimed_total'] += unclaimed_total
                totals['dotace'] += dotace
                totals['final_price'] += final_price
        
        else:  # grouping == 'total' - seskupení podle zákazníka + druhu + jídla
            grouped = defaultdict(
                lambda: {
                    'quantity': 0,
                    'unclaimed_total': 0,
                    'dotace': 0,
                    'final_price': 0,
                }
            )
            
            for item in queryset:
                jidlo_nazev = getattr(item.menu_item.jidlo, 'nazev', 'N/A')
                jidlo_cena = getattr(item.menu_item.jidlo, 'cena', 0)
                druh_jidla = getattr(item.menu_item.druh_jidla, 'nazev', 'Neurčeno')
                user_name = item.order.user.get_full_name()
                
                # Klíč: zákazník + druh jídla + název jídla
                key = (user_name, druh_jidla, jidlo_nazev)
                
                unclaimed_total = item.quantity * jidlo_cena
                dotace = (jidlo_cena - item.cena) * item.quantity
                final_price = item.quantity * item.cena
                
                grouped[key]['quantity'] += item.quantity
                grouped[key]['unclaimed_total'] += unclaimed_total
                grouped[key]['dotace'] += dotace
                grouped[key]['final_price'] += final_price
            
            # Převeď na list
            for (user_name, druh_jidla, jidlo_nazev), data in sorted(
                grouped.items(), 
                key=lambda x: x[1]['final_price'], 
                reverse=True
            ):
                row = {
                    'user': user_name,
                    'food_type': druh_jidla,
                    'food_name': jidlo_nazev,
                    'quantity': data['quantity'],
                    'unclaimed_total': round(data['unclaimed_total'], 2),
                    'dotace': round(data['dotace'], 2),
                    'final_price': round(data['final_price'], 2),
                    'grouped': True
                }
                report_data.append(row)
                
                totals['total_portions'] += data['quantity']
                totals['unclaimed_total'] += data['unclaimed_total']
                totals['dotace'] += data['dotace']
                totals['final_price'] += data['final_price']

        # Fulltext search
        search_query = form.cleaned_data.get('search', '').strip()
        if search_query:
            filtered_data = []
            for row in report_data:
                searchable_text = ' '.join([
                    str(row.get('date', '')),
                    str(row.get('user', '')),
                    str(row.get('food_type', '')),
                    str(row.get('food_name', '')),
                    str(row.get('quantity', '')),
                    str(row.get('unclaimed_total', '')),
                    str(row.get('dotace', '')),
                    str(row.get('final_price', '')),
                ]).lower()
                
                if search_query.lower() in searchable_text:
                    filtered_data.append(row)
            
            report_data = filtered_data
        
        totals['rows_count'] = len(report_data)
        return report_data, totals, grouping

    def get_carkovnice_report(self, form):
        """Měsíční čárkovnice čerpání stravy podle dnů a druhů jídel."""
        if not form.is_valid():
            return [], {}, "month"

        today = timezone.localdate()
        selected_month = form.cleaned_data.get('month') or today.month
        selected_year = form.cleaned_data.get('year') or today.year
        _, days_in_month = calendar.monthrange(selected_year, selected_month)
        date_from = today.replace(year=selected_year, month=selected_month, day=1)
        date_to = today.replace(year=selected_year, month=selected_month, day=days_in_month)

        queryset = (
            OrderItem.objects.filter(
                order__datum_vydeje__range=[date_from, date_to],
                order__status__in=['vydano', 'castecne-vydano'],
            )
            .filter(Q(vydano=True) | Q(order__status='vydano'))
            .select_related('order__user', 'menu_item__druh_jidla')
            .order_by('order__user__last_name', 'order__user__first_name', 'order__datum_vydeje', 'menu_item__druh_jidla__poradi')
        )

        if form.cleaned_data['group']:
            queryset = queryset.filter(order__user__groups=form.cleaned_data['group'])

        if form.cleaned_data['customer']:
            queryset = queryset.filter(order__user=form.cleaned_data['customer'])

        selected_food_types = list(form.cleaned_data.get('food_types') or [])
        if selected_food_types:
            queryset = queryset.filter(menu_item__druh_jidla__in=selected_food_types)

        search_query = form.cleaned_data.get('search', '').strip().lower()
        rows_by_user = {}
        food_type_names = {}
        financial_by_type = defaultdict(
            lambda: {
                'count': 0,
                'subsidized_count': 0,
                'full_price_total': Decimal('0.00'),
                'subsidy_total': Decimal('0.00'),
                'paid_total': Decimal('0.00'),
            }
        )
        finance_totals = {
            'full_price_total': Decimal('0.00'),
            'subsidy_total': Decimal('0.00'),
            'paid_total': Decimal('0.00'),
            'subsidized_portions': 0,
        }
        day_totals = {
            day: {'portions': 0, 'users': set()}
            for day in range(1, days_in_month + 1)
        }

        for item in queryset:
            user = item.order.user
            full_name = user.get_full_name() or user.username
            searchable = " ".join(
                [
                    full_name,
                    user.osobni_cislo or "",
                    user.identifikacni_medium or "",
                ]
            ).lower()
            if search_query and search_query not in searchable:
                continue

            user_row = rows_by_user.setdefault(
                user.pk,
                {
                    'user': full_name,
                    'osobni_cislo': user.osobni_cislo or '',
                    'identifikacni_medium': user.identifikacni_medium or '',
                    'days': {day: [] for day in range(1, days_in_month + 1)},
                    'totals_by_type': defaultdict(int),
                    'financial_by_type': defaultdict(
                        lambda: {
                            'count': 0,
                            'subsidized_count': 0,
                            'full_price_total': Decimal('0.00'),
                            'subsidy_total': Decimal('0.00'),
                            'paid_total': Decimal('0.00'),
                        }
                    ),
                    'total_portions': 0,
                },
            )

            food_type = item.menu_item.druh_jidla
            food_type_names[food_type.pk] = food_type.nazev
            label = food_type_short_label(food_type.nazev)
            full_price_unit = Decimal(str(getattr(item.menu_item.jidlo, 'cena', 0) or 0))
            paid_price_unit = Decimal(str(item.cena or 0))
            subsidy_unit = max(Decimal('0.00'), full_price_unit - paid_price_unit)
            full_price_total = full_price_unit * item.quantity
            subsidy_total = subsidy_unit * item.quantity
            paid_total = paid_price_unit * item.quantity
            subsidized_count = item.quantity if subsidy_unit > 0 else 0

            day = item.order.datum_vydeje.day
            for _ in range(item.quantity):
                user_row['days'][day].append(label)
            user_row['totals_by_type'][food_type.nazev] += item.quantity
            user_row['financial_by_type'][food_type.nazev]['count'] += item.quantity
            user_row['financial_by_type'][food_type.nazev]['subsidized_count'] += subsidized_count
            user_row['financial_by_type'][food_type.nazev]['full_price_total'] += full_price_total
            user_row['financial_by_type'][food_type.nazev]['subsidy_total'] += subsidy_total
            user_row['financial_by_type'][food_type.nazev]['paid_total'] += paid_total
            user_row['total_portions'] += item.quantity

            financial_by_type[food_type.nazev]['count'] += item.quantity
            financial_by_type[food_type.nazev]['subsidized_count'] += subsidized_count
            financial_by_type[food_type.nazev]['full_price_total'] += full_price_total
            financial_by_type[food_type.nazev]['subsidy_total'] += subsidy_total
            financial_by_type[food_type.nazev]['paid_total'] += paid_total

            finance_totals['full_price_total'] += full_price_total
            finance_totals['subsidy_total'] += subsidy_total
            finance_totals['paid_total'] += paid_total
            finance_totals['subsidized_portions'] += subsidized_count
            day_totals[day]['portions'] += item.quantity
            day_totals[day]['users'].add(user.pk)

        ordered_food_types = []
        if selected_food_types:
            ordered_food_types = list(selected_food_types)
        elif food_type_names:
            ordered_food_types = list(
                DruhJidla.objects.filter(pk__in=food_type_names.keys()).order_by('poradi', 'nazev')
            )

        rows = []
        for row in rows_by_user.values():
            row['day_strings'] = {
                day: " ".join(labels)
                for day, labels in row['days'].items()
            }
            row['summary_strings'] = [
                f"{food_type_short_label(food_type.nazev)} {row['totals_by_type'].get(food_type.nazev, 0)}x"
                for food_type in ordered_food_types
                if row['totals_by_type'].get(food_type.nazev, 0)
            ]
            rows.append(row)

        rows.sort(key=lambda entry: (entry['user'].split()[-1].lower(), entry['user'].lower()))

        legend = [
            {
                'name': food_type.nazev,
                'label': food_type_short_label(food_type.nazev),
            }
            for food_type in ordered_food_types
        ]

        financial_rows = [
            {
                'food_type': food_type.nazev,
                'label': food_type_short_label(food_type.nazev),
                'count': financial_by_type[food_type.nazev]['count'],
                'subsidized_count': financial_by_type[food_type.nazev]['subsidized_count'],
                'full_price_total': financial_by_type[food_type.nazev]['full_price_total'],
                'subsidy_total': financial_by_type[food_type.nazev]['subsidy_total'],
                'paid_total': financial_by_type[food_type.nazev]['paid_total'],
            }
            for food_type in ordered_food_types
            if financial_by_type[food_type.nazev]['count']
        ]

        totals = {
            'days': list(range(1, days_in_month + 1)),
            'month_label': f"{selected_month:02d}/{selected_year}",
            'month_name': dict(MONTH_CHOICES).get(selected_month, selected_month),
            'year': selected_year,
            'days_in_month': days_in_month,
            'users_count': len(rows),
            'total_portions': sum(row['total_portions'] for row in rows),
            'subsidized_portions': finance_totals['subsidized_portions'],
            'full_price_total': finance_totals['full_price_total'],
            'subsidy_total': finance_totals['subsidy_total'],
            'paid_total': finance_totals['paid_total'],
            'legend': legend,
            'financial_rows': financial_rows,
            'day_totals': {
                day: {
                    'portions': values['portions'],
                    'users': len(values['users']),
                }
                for day, values in day_totals.items()
            },
            'peak_day': max(
                (
                    {
                        'day': day,
                        'portions': values['portions'],
                        'users': len(values['users']),
                    }
                    for day, values in day_totals.items()
                ),
                key=lambda item: item['portions'],
                default={'day': None, 'portions': 0, 'users': 0},
            ),
        }
        return rows, totals, 'month'

    def get_financial_subsidy_report(self, form):
        """Finanční přehled dotovaných objednávek pro účetní oddělení."""
        report_data = []
        totals = {
            'total_portions': 0,
            'subsidized_portions': 0,
            'full_price_total': Decimal('0.00'),
            'subsidy_total': Decimal('0.00'),
            'paid_total': Decimal('0.00'),
        }
        grouping = 'day'

        if not form.is_valid():
            return report_data, totals, grouping

        grouping = form.cleaned_data.get('grouping', 'day')
        date_from, date_to = self.resolve_period_range(form)

        queryset = (
            OrderItem.objects.filter(
                order__status__in=['vydano', 'castecne-vydano', 'nevyzvednuto'],
            )
            .select_related('order__user', 'menu_item__jidlo', 'menu_item__druh_jidla')
            .order_by('order__datum_vydeje', 'order__user__last_name', 'order__user__first_name')
        )

        if date_from and date_to:
            queryset = queryset.filter(order__datum_vydeje__range=[date_from, date_to])

        if form.cleaned_data['group']:
            queryset = queryset.filter(order__user__groups=form.cleaned_data['group'])

        if form.cleaned_data['customer']:
            queryset = queryset.filter(order__user=form.cleaned_data['customer'])

        if form.cleaned_data.get('food_types'):
            queryset = queryset.filter(menu_item__druh_jidla__in=form.cleaned_data['food_types'])

        search_query = form.cleaned_data.get('search', '').strip().lower()

        for item in queryset:
            user = item.order.user
            full_price_unit = Decimal(str(getattr(item.menu_item.jidlo, 'cena', 0) or 0))
            paid_price_unit = Decimal(str(item.cena or 0))
            subsidy_unit = max(Decimal('0.00'), full_price_unit - paid_price_unit)

            full_price_total = full_price_unit * item.quantity
            subsidy_total = subsidy_unit * item.quantity
            paid_total = paid_price_unit * item.quantity
            subsidized_portions = item.quantity if subsidy_unit > 0 else 0

            row = {
                'user': user.get_full_name() or user.username,
                'osobni_cislo': user.osobni_cislo or '',
                'identifikacni_medium': user.identifikacni_medium or '',
                'date': item.order.datum_vydeje,
                'status': item.order.status,
                'food_type': getattr(item.menu_item.druh_jidla, 'nazev', 'Neurčeno'),
                'food_name': getattr(item.menu_item.jidlo, 'nazev', 'N/A'),
                'quantity': item.quantity,
                'subsidized_portions': subsidized_portions,
                'full_price_total': full_price_total,
                'subsidy_total': subsidy_total,
                'paid_total': paid_total,
            }

            searchable_text = ' '.join([
                row['user'],
                row['osobni_cislo'],
                row['identifikacni_medium'],
                row['food_type'],
                row['food_name'],
                str(row['date']),
                row['status'],
            ]).lower()
            if search_query and search_query not in searchable_text:
                continue

            report_data.append(row)
            totals['total_portions'] += item.quantity
            totals['subsidized_portions'] += subsidized_portions
            totals['full_price_total'] += full_price_total
            totals['subsidy_total'] += subsidy_total
            totals['paid_total'] += paid_total

        if grouping == 'total':
            grouped = defaultdict(
                lambda: {
                    'full_price_total': Decimal('0.00'),
                    'subsidy_total': Decimal('0.00'),
                    'paid_total': Decimal('0.00'),
                    'total_portions': 0,
                    'subsidized_portions': 0,
                    'food_types': set(),
                    'osobni_cislo': '',
                    'identifikacni_medium': '',
                }
            )

            for row in report_data:
                key = row['user']
                grouped[key]['full_price_total'] += row['full_price_total']
                grouped[key]['subsidy_total'] += row['subsidy_total']
                grouped[key]['paid_total'] += row['paid_total']
                grouped[key]['total_portions'] += row['quantity']
                grouped[key]['subsidized_portions'] += row['subsidized_portions']
                grouped[key]['food_types'].add(row['food_type'])
                if not grouped[key]['osobni_cislo']:
                    grouped[key]['osobni_cislo'] = row['osobni_cislo']
                if not grouped[key]['identifikacni_medium']:
                    grouped[key]['identifikacni_medium'] = row['identifikacni_medium']

            report_data = [
                {
                    'user': user_name,
                    'osobni_cislo': data['osobni_cislo'],
                    'identifikacni_medium': data['identifikacni_medium'],
                    'food_types': ', '.join(sorted(data['food_types'])),
                    'total_portions': data['total_portions'],
                    'subsidized_portions': data['subsidized_portions'],
                    'full_price_total': data['full_price_total'],
                    'subsidy_total': data['subsidy_total'],
                    'paid_total': data['paid_total'],
                    'grouped': True,
                }
                for user_name, data in sorted(grouped.items(), key=lambda item: item[0].lower())
            ]

        return report_data, totals, grouping


    def dashboard_view(self, request):
        active_report = request.GET.get('report', 'castky')
        export_type = request.GET.get('export')
        form_data = request.GET.copy() if request.GET else None
        if active_report == 'carkovnice':
            previous_month, previous_year = get_previous_month_year()
            if form_data is None:
                form_data = {}
            form_data.setdefault('period', 'current_month')
            form_data.setdefault('grouping', 'total')
            form_data.setdefault('month', str(previous_month))
            form_data.setdefault('year', str(previous_year))
        form = ReportForm(form_data or None)
        
        # Rozhodnutí, který report použít
        if active_report == 'polozky':
            report_data, totals, grouping = self.get_order_items_report(form)
            report_type = 'items'
        elif active_report == 'finance-dotace':
            report_data, totals, grouping = self.get_financial_subsidy_report(form)
            report_type = 'subsidy_finance'
        elif active_report == 'carkovnice':
            report_data, totals, grouping = self.get_carkovnice_report(form)
            report_type = 'attendance_matrix'
        elif active_report == 'druhy-jidel':  # NOVÉ
            report_data, totals, grouping = self.get_food_types_report(form)
            report_type = 'food_types'
        else:
            report_data, totals, grouping = self.get_report_calculations(form)
            report_type = 'amounts'

        if export_type and report_data:
            filename_base = f"report_{active_report}_{timezone.now().strftime('%Y%m%d')}"

            # CSV Export
            if export_type == 'csv':
                response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
                writer = csv.writer(response, delimiter=';')
                
                if report_type == 'subsidy_finance':
                    headers = ['Zákazník', 'Osobní číslo', 'ID médium']
                    if grouping == 'day':
                        headers += ['Datum', 'Stav', 'Druh jídla', 'Název jídla', 'Porcí', 'Dotovaných porcí']
                    else:
                        headers += ['Druhy jídel', 'Porcí', 'Dotovaných porcí']
                    headers += ['Nedotovaná cena', 'Dotace', 'Cena po dotaci']
                elif report_type == 'attendance_matrix':
                    headers = ['Zákazník', 'Osobní číslo', 'ID médium'] + [str(day) for day in totals['days']] + ['Souhrn', 'Celkem porcí']
                elif report_type == 'food_types':
                    headers = []
                    if grouping == 'day': headers += ['Datum']
                    headers += ['Zákazník', 'Druh jídla', 'Název jídla', 'Počet ks', 'Plná částka', 'Dotace', 'K platbě']
                else:
                    headers = ['Zákazník', 'Osobní číslo', 'ID médium']
                    if grouping == 'day': headers += ['Datum', 'Stav']
                    if report_type == 'items':
                        headers += (['Jídlo', 'Počet ks'] if grouping == 'day' else ['Počet položek', 'Jídla'])
                    headers += ['Plná částka', 'Dotace', 'K platbě']
                    if report_type == 'amounts' and grouping == 'total': headers += ['Počet objednávek']
                
                writer.writerow(headers)
                
                for row in report_data:
                    if report_type == 'subsidy_finance':
                        line = [row['user'], row['osobni_cislo'], row['identifikacni_medium']]
                        if grouping == 'day':
                            line += [
                                row['date'].strftime('%d.%m.%Y'),
                                row['status'],
                                row['food_type'],
                                row['food_name'],
                                row['quantity'],
                                row['subsidized_portions'],
                            ]
                        else:
                            line += [
                                row.get('food_types', ''),
                                row['total_portions'],
                                row['subsidized_portions'],
                            ]
                        line += [
                            fmt_max2(row['full_price_total']).replace('.', ','),
                            fmt_max2(row['subsidy_total']).replace('.', ','),
                            fmt_max2(row['paid_total']).replace('.', ','),
                        ]
                    elif report_type == 'attendance_matrix':
                        line = [
                            row['user'],
                            row['osobni_cislo'],
                            row['identifikacni_medium'],
                        ] + [
                            row['day_strings'].get(day, '')
                            for day in totals['days']
                        ] + [
                            ', '.join(row['summary_strings']),
                            row['total_portions'],
                        ]
                    elif report_type == 'food_types':
                        line = []
                        if grouping == 'day': line += [row['date'].strftime('%d.%m.%Y')]
                        line += [row['user'], row['food_type'], row['food_name'], row['quantity'],
                                fmt_max2(row['unclaimed_total']).replace('.', ','),
                                fmt_max2(row['dotace']).replace('.', ','),
                                fmt_max2(row['final_price']).replace('.', ',')]
                    else:
                        line = [row['user'], row['osobni_cislo'], row['identifikacni_medium']]
                        if grouping == 'day': line += [row['date'].strftime('%d.%m.%Y'), row['status']]
                        if report_type == 'items':
                            line += ([row.get('jidlo_nazev', ''), row.get('quantity', 0)] if grouping == 'day' 
                                    else [row.get('items_count', 0), row.get('items_names', '')])
                        line += [fmt_max2(row['unclaimed_total']).replace('.', ','),
                                fmt_max2(row['dotace']).replace('.', ','),
                                fmt_max2(row['final_price']).replace('.', ',')]
                        if report_type == 'amounts' and grouping == 'total': line += [row.get('count', 0)]
                    writer.writerow(line)
                return response

            # XLSX Export
            elif export_type == 'xls' and OPENPYXL_AVAILABLE:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Report"
                
                if report_type == 'subsidy_finance':
                    headers = ['Zákazník', 'Osobní číslo', 'ID médium']
                    if grouping == 'day':
                        headers += ['Datum', 'Stav', 'Druh jídla', 'Název jídla', 'Porcí', 'Dotovaných porcí']
                    else:
                        headers += ['Druhy jídel', 'Porcí', 'Dotovaných porcí']
                    headers += ['Nedotovaná cena', 'Dotace', 'Cena po dotaci']
                elif report_type == 'attendance_matrix':
                    headers = ['Zákazník', 'Osobní číslo', 'ID médium'] + [str(day) for day in totals['days']] + ['Souhrn', 'Celkem porcí']
                elif report_type == 'food_types':
                    headers = []
                    if grouping == 'day': headers += ['Datum']
                    headers += ['Zákazník', 'Druh jídla', 'Název jídla', 'Počet ks', 'Plná částka', 'Dotace', 'K platbě']
                else:
                    headers = ['Zákazník', 'Osobní číslo', 'ID médium']
                    if grouping == 'day': headers += ['Datum', 'Stav']
                    if report_type == 'items':
                        headers += (['Jídlo', 'Počet ks'] if grouping == 'day' else ['Počet položek', 'Jídla'])
                    headers += ['Plná částka', 'Dotace', 'K platbě']
                    if report_type == 'amounts' and grouping == 'total': headers += ['Počet objednávek']
                
                ws.append(headers)
                
                for row in report_data:
                    if report_type == 'subsidy_finance':
                        line = [row['user'], row['osobni_cislo'], row['identifikacni_medium']]
                        if grouping == 'day':
                            line += [
                                row['date'].strftime('%d.%m.%Y'),
                                row['status'],
                                row['food_type'],
                                row['food_name'],
                                row['quantity'],
                                row['subsidized_portions'],
                            ]
                        else:
                            line += [
                                row.get('food_types', ''),
                                row['total_portions'],
                                row['subsidized_portions'],
                            ]
                        line += [
                            fmt_max2(row['full_price_total']),
                            fmt_max2(row['subsidy_total']),
                            fmt_max2(row['paid_total']),
                        ]
                    elif report_type == 'attendance_matrix':
                        line = [
                            row['user'],
                            row['osobni_cislo'],
                            row['identifikacni_medium'],
                        ] + [
                            row['day_strings'].get(day, '')
                            for day in totals['days']
                        ] + [
                            ', '.join(row['summary_strings']),
                            row['total_portions'],
                        ]
                    elif report_type == 'food_types':
                        line = []
                        if grouping == 'day': line += [row['date'].strftime('%d.%m.%Y')]
                        line += [row['user'], row['food_type'], row['food_name'], row['quantity'],
                                fmt_max2(row['unclaimed_total']), fmt_max2(row['dotace']), fmt_max2(row['final_price'])]
                    else:
                        line = [row['user'], row['osobni_cislo'], row['identifikacni_medium']]
                        if grouping == 'day': line += [row['date'].strftime('%d.%m.%Y'), row['status']]
                        if report_type == 'items':
                            line += ([row.get('jidlo_nazev', ''), row.get('quantity', 0)] if grouping == 'day' 
                                    else [row.get('items_count', 0), row.get('items_names', '')])
                        line += [fmt_max2(row['unclaimed_total']), fmt_max2(row['dotace']), fmt_max2(row['final_price'])]
                        if report_type == 'amounts' and grouping == 'total': line += [row.get('count', 0)]
                    ws.append(line)
                
                # Footer XLSX
                footer = ['CELKEM']
                if report_type == 'subsidy_finance':
                    footer += [''] * (len(headers) - 4)
                    footer += [
                        totals['full_price_total'],
                        totals['subsidy_total'],
                        totals['paid_total'],
                    ]
                elif report_type == 'attendance_matrix':
                    footer += [''] * (len(headers) - 2)
                    footer += [totals['total_portions']]
                else:
                    col_offset = len(headers) - 3
                    footer += [''] * (col_offset - 1)
                    footer += [totals.get('total_portions', totals.get('total_items', '')),
                              fmt_max2(totals['unclaimed_total']), fmt_max2(totals['dotace']), fmt_max2(totals['final_price'])]
                ws.append(footer)
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
                
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
                wb.save(response)
                return response

            # PDF Export
            elif export_type == 'pdf':
                return self.build_report_pdf(report_type, grouping, form, report_data, totals, filename_base)



        report_titles = {
            'amounts': 'Částky objednávek',
            'items': 'Položky objednávek',
            'food_types': 'Druhy jídel',
            'attendance_matrix': 'Čárkovnice',
            'subsidy_finance': 'Finance a dotace',
        }
        report_descriptions = {
            'amounts': 'Souhrnný přehled objednávek, dotací a ceny k platbě podle období a zákazníků.',
            'items': 'Detailní rozpad jednotlivých vydaných jídel a jejich cen.',
            'food_types': 'Přehled čerpání a cen po druzích jídel a konkrétních pokrmech.',
            'attendance_matrix': 'Měsíční čárkovnice skutečně vydané stravy s finančním souhrnem po druzích jídel.',
            'subsidy_finance': 'Finanční podklad pro účetní oddělení s plnou cenou, dotací a konečnou cenou.',
        }
        selected_group = form.cleaned_data.get('group') if form.is_valid() else None
        selected_customer = form.cleaned_data.get('customer') if form.is_valid() else None
        selected_food_types = list(form.cleaned_data.get('food_types') or []) if form.is_valid() else []
        search_query = form.cleaned_data.get('search', '').strip() if form.is_valid() else ''
        if report_type == 'attendance_matrix' and form.is_valid():
            period_label = f"{dict(MONTH_CHOICES).get(form.cleaned_data.get('month'), form.cleaned_data.get('month'))} {form.cleaned_data.get('year')}"
        else:
            period_label = report_period_label(form)

        context = {
            **self.admin_site.each_context(request),
            'form': form,
            'report_data': report_data,
            'totals': totals,
            'active_report': active_report,
            'report_type': report_type,
            'grouping': grouping,
            'report_title': report_titles.get(report_type, 'Report'),
            'report_description': report_descriptions.get(report_type, ''),
            'period_label': period_label,
            'selected_group_label': selected_group.name if selected_group else 'Všechny skupiny',
            'selected_customer_label': (selected_customer.get_full_name() or selected_customer.username) if selected_customer else 'Všichni zákazníci',
            'selected_food_types_label': ', '.join(food_type.nazev for food_type in selected_food_types) if selected_food_types else 'Všechny druhy jídel',
            'search_query': search_query,
            'reports': [
                {'id': 'castky', 'title': 'Částky objednávek', 'icon': 'fas fa-wallet'},
                {'id': 'finance-dotace', 'title': 'Finance a dotace', 'icon': 'fas fa-coins'},
                {'id': 'carkovnice', 'title': 'Čárkovnice', 'icon': 'fas fa-table-cells-large'},
                {'id': 'druhy-jidel', 'title': 'Druhy jídel', 'icon': 'fas fa-layer-group'},
            ],
            'opts': self.model._meta,
            'title': 'Reporty Dashboard',
        }

        return render(request, 'admin/reporty/dashboard.html', context)
