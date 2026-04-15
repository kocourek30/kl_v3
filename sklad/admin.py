from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
import calendar
import os

from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils.dateparse import parse_date
from django.utils.html import escape, format_html
from django.utils.safestring import mark_safe
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)

from django.contrib.admin import ModelAdmin

from objednavky.models import OrderItem
from users.models import StravovaciSkupina
from .models import (
    SkladDashboard,
    Surovina,
    StavSkladu,
    Dodavatel,
    PohybSkladu,
    RecepturaPolozka,
    KomponentaJidla,
    KomponentaSurovina,
    JidloKomponenta,
    PrijemSkladu,
    PolozkaPrijmu,
    Inventura,
    PolozkaInventury,
    PolozkaInventurySarze,
    InventurniDoklad,
    Vydejka,
    PolozkaVydejky,
    NormaSpotrebnihoKose,
    ReportNakladySkladu,
    ToleranceSpotrebnihoKose,
    SarzeSkladu,
    OdpisExpirace,
    SkladovaUzaverka,
)

from .services import (
    generate_vydejka_from_orders,
    objednavky_rekap_data,
    get_order_items_for_vydejka,
    najdi_nedostatecne_stavy_pro_vydejku,
    spocitej_spotrebu_jidla,
    uzavri_prijem,
    uzavri_vydejku,
    uzavri_inventuru,
    stornuj_prijem,
    stornuj_vydejku,
    stornuj_inventuru,
    uzavri_odpis_expirace,
    aktualizuj_stavy_sarzi,
    souhrn_odpisu_expirace,
    validace_surovin_pro_sk,
    spocitej_stravnikodny_obdobi,
    spocitej_spotrebu_sk_mesic,
    priprav_radky_spotrebi_kos_tabulka,
    spocitej_naklady_mesic,
    mesicni_skladova_uzaverka,
    denni_skladovy_checklist,
    inventurni_nahled,
    karta_suroviny_data,
    managersky_report_skladu,
    najdi_rozdily_stav_vs_sarze,
    napln_sarzovou_inventuru,
    nahled_vydejky,
    navrh_nakupu,
    pruvodce_skladovou_uzaverkou,
    souhrn_sarzove_inventury,
    synchronizuj_surovinove_polozky_inventury,
    uzavri_skladovou_uzaverku,
    otevri_skladovou_uzaverku,
    validace_prijemky_pred_uzavrenim,
    format_cena_za_jednotku,
    format_mnozstvi_s_jednotkou,
    priprav_naklady_podle_skupin_sk,
    spocitej_podil_masnych_vyrobku,
    spocitej_podil_bio,
    spocitej_volny_cukr,
    spocitej_legislativni_ukazatele_sk,
    spocitej_souhrn_spotrebniho_kose,
    zkontroluj_jidelnicek_sk,
    zdravi_skladu,
)

from .forms import SpotrebniKosForm


def _prepare_uzavreni_po_ulozeni(model, obj, change):
    """
    Admin nejdřív ukládá model a až potom inline položky.
    Pokud uživatel zaškrtne uzavření, necháme doklad dočasně otevřený
    a skladovou službu zavoláme až po uložení všech položek.
    """
    puvodni_uzavreny = False
    if change and obj.pk:
        puvodni_uzavreny = bool(
            model.objects
            .filter(pk=obj.pk)
            .values_list("uzavreny", flat=True)
            .first()
        )

    uzavrit_po_ulozeni = bool(obj.uzavreny and not puvodni_uzavreny)
    obj._puvodni_uzavreny = puvodni_uzavreny
    obj._uzavrit_po_ulozeni = uzavrit_po_ulozeni

    if uzavrit_po_ulozeni:
        obj.uzavreny = False
        if hasattr(obj, "uzavren_at"):
            obj.uzavren_at = None
        if hasattr(obj, "uzavrel"):
            obj.uzavrel = None

    return obj


def _pdf_styles():
    styles = getSampleStyleSheet()
    font_name = "Helvetica"
    font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "DejaVuSans.ttf")
    if os.path.exists(font_path):
        try:
            if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
            font_name = "DejaVuSans"
        except Exception:
            font_name = "Helvetica"

    for style_name in ("Normal", "Title", "Heading2"):
        styles[style_name].fontName = font_name
    return styles, font_name


def _inventura_pdf_data(inventura):
    if inventura.sarze_polozky.exists():
        rows = []
        manko = Decimal("0")
        prebytek = Decimal("0")
        for pol in inventura.sarze_polozky.select_related("surovina", "sarze_skladu").order_by("surovina__nazev", "datum_spotreby", "id"):
            cena = pol.cena_za_jednotku or pol.surovina.prumerna_cena_za_jednotku or Decimal("0")
            rozdil = pol.rozdil or Decimal("0")
            hodnota = abs(rozdil) * cena
            if rozdil < 0:
                manko += hodnota
            elif rozdil > 0:
                prebytek += hodnota
            rows.append([
                pol.surovina.nazev,
                pol.sarze or "-",
                pol.datum_spotreby.strftime("%d.%m.%Y") if pol.datum_spotreby else "-",
                format_mnozstvi_s_jednotkou(pol.surovina, pol.stav_pred),
                format_mnozstvi_s_jednotkou(pol.surovina, pol.fyzicky_stav),
                format_mnozstvi_s_jednotkou(pol.surovina, rozdil),
                f"{hodnota:.2f} Kč",
                pol.poznamka or "",
            ])
        return {"mode": "sarze", "rows": rows, "manko": manko, "prebytek": prebytek, "cisty_rozdil": prebytek - manko}

    rows = []
    manko = Decimal("0")
    prebytek = Decimal("0")
    for pol in inventura.polozky.select_related("surovina").order_by("surovina__nazev", "id"):
        cena = pol.surovina.prumerna_cena_za_jednotku or Decimal("0")
        rozdil = pol.rozdil or Decimal("0")
        hodnota = abs(rozdil) * cena
        if rozdil < 0:
            manko += hodnota
        elif rozdil > 0:
            prebytek += hodnota
        rows.append([
            pol.surovina.nazev,
            "-",
            "-",
            format_mnozstvi_s_jednotkou(pol.surovina, pol.stav_pred),
            format_mnozstvi_s_jednotkou(pol.surovina, pol.fyzicky_stav),
            format_mnozstvi_s_jednotkou(pol.surovina, rozdil),
            f"{hodnota:.2f} Kč",
            "",
        ])
    return {"mode": "suroviny", "rows": rows, "manko": manko, "prebytek": prebytek, "cisty_rozdil": prebytek - manko}


def inventura_pdf_view(request, inventura_id):
    inventura = get_object_or_404(Inventura, pk=inventura_id)
    styles, font_name = _pdf_styles()
    data = _inventura_pdf_data(inventura)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.1 * cm,
        rightMargin=1.1 * cm,
    )

    stav = "uzavřená" if inventura.uzavreny else "otevřená"
    if inventura.stornovano:
        stav = "stornovaná"

    story = [
        Paragraph(f"Inventurní protokol #{inventura.id}", styles["Title"]),
        Spacer(1, 0.25 * cm),
        Paragraph(f"Datum inventury: {inventura.datum.strftime('%d.%m.%Y')}", styles["Normal"]),
        Paragraph(f"Stav dokladu: {stav}", styles["Normal"]),
        Paragraph(f"Vytvořil: {escape(str(inventura.vytvoril or '-'))}", styles["Normal"]),
        Paragraph(f"Uzavřel: {escape(str(inventura.uzavrel or '-'))}", styles["Normal"]),
        Paragraph(f"Poznámka: {escape(inventura.popis or '-')}", styles["Normal"]),
        Spacer(1, 0.4 * cm),
        Paragraph("Souhrn rozdílů", styles["Heading2"]),
    ]

    summary_table = Table(
        [
            ["Manko", f"{data['manko']:.2f} Kč"],
            ["Přebytek", f"{data['prebytek']:.2f} Kč"],
            ["Čistý rozdíl", f"{data['cisty_rozdil']:.2f} Kč"],
            ["Režim", "Šaržová inventura" if data["mode"] == "sarze" else "Surovinová inventura"],
        ],
        colWidths=[6 * cm, 5 * cm],
    )
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
    ]))
    story += [summary_table, Spacer(1, 0.5 * cm), Paragraph("Položky inventury", styles["Heading2"])]

    table_data = [["Surovina", "Šarže", "Spotřeba", "Účetní stav", "Fyzický stav", "Rozdíl", "Hodnota", "Poznámka"]]
    table_data.extend(data["rows"] or [["Bez položek", "-", "-", "-", "-", "-", "-", "-"]])
    table = Table(
        table_data,
        colWidths=[3.4 * cm, 2.2 * cm, 2.0 * cm, 2.3 * cm, 2.3 * cm, 2.0 * cm, 2.1 * cm, 2.0 * cm],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
        ("ALIGN", (3, 1), (6, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(table)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="inventura_{inventura.id}.pdf"'
    return response


def _uzaverka_mesice_link(obj):
    if not obj or not getattr(obj, "datum", None):
        return "-"
    url = reverse("admin:sklad_uzaverka_mesic", args=[obj.datum.year, obj.datum.month])
    return format_html(
        '<a class="button" href="{}">Uzávěrka {:02d}/{}</a>',
        url,
        obj.datum.month,
        obj.datum.year,
    )


def _dopln_vydejku_z_objednavek_pokud_je_prazdna(vydejka):
    if not vydejka.polozky.exists():
        generate_vydejka_from_orders(
            datum=vydejka.datum,
            stravovaci_skupina=vydejka.stravovaci_skupina,
            typ_stravy=vydejka.typ_stravy,
        )


def _upozorni_na_nedostatecne_stavy(request, vydejka):
    nedostatky = najdi_nedostatecne_stavy_pro_vydejku(vydejka)
    if not nedostatky:
        return

    popis = ", ".join(
        f"{radek['surovina'].nazev}: chybí {radek['chybi']} {radek['surovina'].jednotka}"
        for radek in nedostatky[:8]
    )
    if len(nedostatky) > 8:
        popis += f" a dalších {len(nedostatky) - 8}"

    messages.warning(
        request,
        f"Výdejka #{vydejka.id} odepsala některé suroviny do mínusu: {popis}.",
    )


def _stav_dokladu_text(obj):
    if not obj:
        return "Rozpracováno"
    if getattr(obj, "stornovano", False):
        return "Stornováno"
    if getattr(obj, "uzavreny", False):
        return "Uzavřeno a promítnuto do skladu"
    return "Rozpracováno, sklad ještě nebyl změněn"


# -------------------------------------------------------------------
# Formulář pro náklady skladu
# -------------------------------------------------------------------


class MesicniNakladyForm(forms.Form):
    rok = forms.IntegerField(label="Rok")
    mesic = forms.IntegerField(min_value=1, max_value=12, label="Měsíc")
    stravovaci_skupina = forms.ModelChoiceField(
        queryset=StravovaciSkupina.objects.all(),
        required=False,
        label="Stravovací skupina",
        help_text="Volitelné – bez výběru se počítají náklady za všechny skupiny.",
    )


class NavrhNakupuForm(forms.Form):
    date_from = forms.DateField(
        label="Od data",
        initial=date.today,
        widget=forms.DateInput(attrs={"type": "date"}),
    )
    date_to = forms.DateField(
        label="Do data",
        required=False,
        widget=forms.DateInput(attrs={"type": "date"}),
        help_text="Když necháš prázdné, počítá se 7 dní od data Od.",
    )

    def clean(self):
        cleaned = super().clean()
        date_from = cleaned.get("date_from")
        date_to = cleaned.get("date_to")
        if date_from and date_to and date_from > date_to:
            raise forms.ValidationError("Datum Od nesmí být po datu Do.")
        return cleaned


class ImportPrijemXlsxForm(forms.Form):
    soubor = forms.FileField(
        label="Excel soubor",
        help_text="Očekávané sloupce: surovina, mnozstvi, jednotkova_cena, pocet_baleni, mnozstvi_v_baleni, jednotka_baleni, cena_za_baleni_bez_dph, sazba_dph, sarze, datum_spotreby.",
    )


@admin.register(ReportNakladySkladu)
class ReportNakladySkladuAdmin(admin.ModelAdmin):
    """
    Pseudo-model pro zobrazení reportu v adminu.
    """

    def get_queryset(self, request):
        return ReportNakladySkladu.objects.none()

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.is_staff

    def has_delete_permission(self, request, obj=None):
        return False

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "",
                self.admin_site.admin_view(self.report_view),
                name="sklad_report_naklady",
            ),
        ]
        return custom + urls

    def report_view(self, request):
        today = date.today()
        initial = {"rok": today.year, "mesic": today.month}
        form = MesicniNakladyForm(request.GET or None, initial=initial)

        souhrnne_naklady = None
        naklady_podle_sk = None
        skladova_uzaverka = None
        managersky_report = None

        if form.is_valid():
            rok = form.cleaned_data["rok"]
            mesic = form.cleaned_data["mesic"]
            skupina = form.cleaned_data["stravovaci_skupina"]

            souhrnne_naklady = spocitej_naklady_mesic(rok, mesic, skupina)
            skladova_uzaverka = mesicni_skladova_uzaverka(rok, mesic)
            naklady_podle_sk = priprav_naklady_podle_skupin_sk(rok, mesic, skupina)
            managersky_report = managersky_report_skladu(rok, mesic)

        context = dict(
            self.admin_site.each_context(request),
            title="Report nákladů na suroviny",
            form=form,
            souhrnne_naklady=souhrnne_naklady,
            naklady_podle_sk=naklady_podle_sk,
            skladova_uzaverka=skladova_uzaverka,
            managersky_report=managersky_report,
        )
        return TemplateResponse(
            request,
            "admin/sklad/report_naklady.html",
            context,
        )


# -------------------------------------------------------------------
# Normy a tolerance spotřebního koše
# -------------------------------------------------------------------


@admin.register(NormaSpotrebnihoKose)
class NormaSKAdmin(admin.ModelAdmin):
    list_display = ("vekova_kategorie", "typ_jidla", "skupina_sk", "norma_g_den", "stravovaci_skupina")
    list_filter = ("vekova_kategorie", "typ_jidla", "skupina_sk", "stravovaci_skupina")
    search_fields = ("stravovaci_skupina__nazev",)


@admin.register(ToleranceSpotrebnihoKose)
class ToleranceSKAdmin(admin.ModelAdmin):
    list_display = ("stravovaci_skupina", "skupina_sk", "min_pct", "max_pct")
    list_filter = ("stravovaci_skupina", "skupina_sk")
    search_fields = ("stravovaci_skupina__nazev",)


# -------------------------------------------------------------------
# Starší měsíční SK formulář (jen měsíc/rok)
# -------------------------------------------------------------------


class MesicniSKForm(forms.Form):
    rok = forms.IntegerField(initial=date.today().year, label="Rok")
    mesic = forms.IntegerField(
        initial=date.today().month, min_value=1, max_value=12, label="Měsíc"
    )
    stravovaci_skupina = forms.ModelChoiceField(
        queryset=StravovaciSkupina.objects.all(),
        required=False,
        label="Stravovací skupina",
        help_text="Vyber skupinu strávníků. Bez výběru se tabulka nepočítá.",
    )


def spocitej_stravnikodny_mesic(rok: int, mesic: int, stravovaci_skupina=None):
    """
    Celkový počet 'strávníkoden' za měsíc = součet porcí.
    """
    qs = OrderItem.objects.filter(
        order__datum_vydeje__year=rok,
        order__datum_vydeje__month=mesic,
    )

    if stravovaci_skupina is not None:
        qs = qs.filter(order__user__stravovaci_skupina=stravovaci_skupina)

    total = qs.aggregate(celkem=Sum("quantity"))["celkem"] or 0
    return Decimal(total)


# -------------------------------------------------------------------
# Nový přehled spotřebního koše (perioda z formuláře SpotrebniKosForm)
# -------------------------------------------------------------------


@admin.register(SkladDashboard)
class SkladSpotrebniKosAdmin(ModelAdmin):
    change_list_template = "admin/sklad/dashboard.html"
    expirace_varovani_dnu = 14
    limit_minima_nasobek = Decimal("1.20")

    def get_queryset(self, request):
        return SkladDashboard.objects.none()

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.is_staff

    def has_delete_permission(self, request, obj=None):
        return False

    def _dashboard_expirace(self, target_date):
        expirace_do = target_date + timedelta(days=self.expirace_varovani_dnu)
        aktualizuj_stavy_sarzi(dnes=target_date)
        sarze_qs = (
            SarzeSkladu.objects
            .filter(
                datum_spotreby__isnull=False,
                datum_spotreby__lte=expirace_do,
                mnozstvi_zbyva__gt=0,
                stav__in=[
                    SarzeSkladu.STAV_POUZITELNA,
                    SarzeSkladu.STAV_KARANTENA,
                    SarzeSkladu.STAV_EXPIROVANA,
                ],
            )
            .select_related("surovina", "polozka_prijmu", "polozka_prijmu__prijem", "polozka_prijmu__prijem__dodavatel")
            .order_by("datum_spotreby", "surovina__nazev", "id")[:20]
        )
        return [
            {
                "surovina": sarze.surovina,
                "sarze": sarze.sarze,
                "stav": sarze.stav,
                "datum_spotreby": sarze.datum_spotreby,
                "mnozstvi_zbyva": sarze.mnozstvi_zbyva,
                "mnozstvi_zbyva_display": format_mnozstvi_s_jednotkou(sarze.surovina, sarze.mnozstvi_zbyva),
                "polozka_prijmu": sarze.polozka_prijmu,
            }
            for sarze in sarze_qs
        ]

    def _dashboard_minimum_alerty(self):
        alerty = []
        stavy = StavSkladu.objects.select_related("surovina").filter(min_mnozstvi__gt=0)
        for stav in stavy:
            mnozstvi = stav.mnozstvi or Decimal("0")
            minimum = stav.min_mnozstvi or Decimal("0")
            limit_blizko = minimum * self.limit_minima_nasobek
            if mnozstvi > limit_blizko:
                continue

            alerty.append({
                "surovina": stav.surovina,
                "mnozstvi": mnozstvi,
                "mnozstvi_display": format_mnozstvi_s_jednotkou(stav.surovina, mnozstvi),
                "minimum": minimum,
                "minimum_display": format_mnozstvi_s_jednotkou(stav.surovina, minimum),
                "limit_blizko": limit_blizko,
                "pod_min": mnozstvi <= minimum,
                "procento_minima": (mnozstvi / minimum * Decimal("100")) if minimum else Decimal("0"),
            })

        return sorted(
            alerty,
            key=lambda row: (not row["pod_min"], row["procento_minima"], row["surovina"].nazev),
        )[:20]

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "mesicni-spotrebni-kos/",
                self.admin_site.admin_view(self.spotrebni_kos_view),
                name="sklad_mesicni_spotrebni_kos",
            ),
            path(
                "mesicni-spotrebni-kos/xls/",
                self.admin_site.admin_view(self.spotrebni_kos_xls_view),
                name="sklad_mesicni_spotrebni_kos_xls",
            ),
            path(
                "mesicni-spotrebni-kos/pdf/",
                self.admin_site.admin_view(self.spotrebni_kos_pdf_view),
                name="sklad_mesicni_spotrebni_kos_pdf",
            ),
            path(
                "navrh-nakupu/",
                self.admin_site.admin_view(self.navrh_nakupu_view),
                name="sklad_navrh_nakupu",
            ),
            path(
                "zdravi-skladu/",
                self.admin_site.admin_view(self.zdravi_skladu_view),
                name="sklad_zdravi_skladu",
            ),
            path(
                "zdravi-skladu/pdf/",
                self.admin_site.admin_view(self.zdravi_skladu_pdf_view),
                name="sklad_zdravi_skladu_pdf",
            ),
        ]
        return custom + urls

    def changelist_view(self, request, extra_context=None):
        target_date = parse_date(request.GET.get("date") or "") or date.today()

        expected = {}
        order_items = (
            OrderItem.objects
            .select_related("menu_item__jidlo")
            .filter(order__datum_vydeje=target_date)
        )
        for item in order_items:
            spotreba = spocitej_spotrebu_jidla(item.menu_item.jidlo, Decimal(item.quantity))
            for surovina_id, mnozstvi in spotreba.items():
                expected[surovina_id] = expected.get(surovina_id, Decimal("0")) + mnozstvi

        real = {}
        pohyby = (
            PohybSkladu.objects
            .filter(vydejka__datum=target_date, typ=PohybSkladu.TYP_VYDEJ)
            .select_related("surovina")
        )
        for pohyb in pohyby:
            real[pohyb.surovina_id] = real.get(pohyb.surovina_id, Decimal("0")) + (pohyb.mnozstvi or Decimal("0"))

        surovina_ids = set(expected.keys()) | set(real.keys())
        suroviny = {
            s.id: s
            for s in Surovina.objects.filter(id__in=surovina_ids).select_related("stav")
        }

        rows = []
        for surovina_id in sorted(surovina_ids, key=lambda sid: suroviny[sid].nazev if sid in suroviny else ""):
            surovina = suroviny.get(surovina_id)
            if not surovina:
                continue
            stav = getattr(surovina, "stav", None)
            stav_mnozstvi = stav.mnozstvi if stav else None
            min_mnozstvi = stav.min_mnozstvi if stav else None
            rows.append({
                "surovina": surovina,
                "expected": expected.get(surovina_id, Decimal("0")),
                "expected_display": format_mnozstvi_s_jednotkou(surovina, expected.get(surovina_id, Decimal("0"))),
                "real": real.get(surovina_id, Decimal("0")),
                "real_display": format_mnozstvi_s_jednotkou(surovina, real.get(surovina_id, Decimal("0"))),
                "stav": stav_mnozstvi,
                "stav_display": format_mnozstvi_s_jednotkou(surovina, stav_mnozstvi) if stav_mnozstvi is not None else None,
                "min": min_mnozstvi,
                "min_display": format_mnozstvi_s_jednotkou(surovina, min_mnozstvi) if min_mnozstvi is not None else None,
                "pod_min": (
                    stav_mnozstvi is not None
                    and min_mnozstvi is not None
                    and stav_mnozstvi <= min_mnozstvi
                ),
                "blizko_min": (
                    stav_mnozstvi is not None
                    and min_mnozstvi is not None
                    and min_mnozstvi > 0
                    and min_mnozstvi < stav_mnozstvi <= (min_mnozstvi * self.limit_minima_nasobek)
                ),
            })

        extra_context = extra_context or {}
        extra_context.update({
            "target_date": target_date,
            "rows": rows,
            "expirace_do": target_date + timedelta(days=self.expirace_varovani_dnu),
            "expirace_polozky": self._dashboard_expirace(target_date),
            "minimum_alerty": self._dashboard_minimum_alerty(),
            "rozdily_stav_sarze": najdi_rozdily_stav_vs_sarze()[:20],
            "denni_checklist": denni_skladovy_checklist(target_date),
            "navrh_nakupu_url": "navrh-nakupu/",
            "zdravi_skladu": zdravi_skladu(target_date),
            "zdravi_skladu_url": "zdravi-skladu/",
            "zdravi_skladu_pdf_url": f"zdravi-skladu/pdf/?date={target_date.isoformat()}",
        })
        return super().changelist_view(request, extra_context=extra_context)

    def zdravi_skladu_view(self, request):
        target_date = parse_date(request.GET.get("date") or "") or date.today()
        data = zdravi_skladu(target_date)
        context = dict(
            self.admin_site.each_context(request),
            title="Kontrolní report zdraví skladu",
            data=data,
            pdf_url=f"pdf/?date={target_date.isoformat()}",
        )
        return TemplateResponse(request, "admin/sklad/zdravi_skladu.html", context)

    def zdravi_skladu_pdf_view(self, request):
        target_date = parse_date(request.GET.get("date") or "") or date.today()
        data = zdravi_skladu(target_date)
        styles, font_name = _pdf_styles()
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.4 * cm,
            leftMargin=1.4 * cm,
            topMargin=1.4 * cm,
            bottomMargin=1.4 * cm,
        )
        story = [
            Paragraph("Kontrolní report zdraví skladu", styles["Title"]),
            Spacer(1, 0.25 * cm),
            Paragraph(f"Datum kontroly: {target_date.strftime('%d.%m.%Y')}", styles["Normal"]),
            Paragraph(f"Hodnota skladu: {data['hodnota_skladu']:.2f} Kč", styles["Normal"]),
            Paragraph(f"Skóre skladu: {data['skore']} %", styles["Normal"]),
            Spacer(1, 0.4 * cm),
        ]
        table_data = [["Kontrola", "Počet", "Stav", "Poznámka"]]
        for row in data["rizika"]:
            table_data.append([
                row["nazev"],
                row["pocet"],
                "V pořádku" if row["ok"] else "Vyžaduje kontrolu",
                row["popis"],
            ])
        table = Table(table_data, colWidths=[5.3 * cm, 1.7 * cm, 3.2 * cm, 7 * cm], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)
        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="zdravi_skladu_{target_date.isoformat()}.pdf"'
        return response

    def navrh_nakupu_view(self, request):
        form = NavrhNakupuForm(request.GET or None)
        data = None
        if form.is_valid():
            date_from = form.cleaned_data["date_from"]
            date_to = form.cleaned_data["date_to"]
            data = navrh_nakupu(date_from=date_from, date_to=date_to)
        elif not request.GET:
            form = NavrhNakupuForm(initial={"date_from": date.today()})
            data = navrh_nakupu(date_from=date.today())

        context = dict(
            self.admin_site.each_context(request),
            title="Návrh nákupu",
            form=form,
            data=data,
        )
        return TemplateResponse(request, "admin/sklad/navrh_nakupu.html", context)

    @method_decorator(never_cache)
    def spotrebni_kos_view(self, request):
        data = self._spotrebni_kos_data(request)
        context = dict(
            self.admin_site.each_context(request),
            **data,
        )
        return TemplateResponse(
            request,
            "admin/sklad/mesicni_spotrebni_kos.html",
            context,
        )

    def _spotrebni_kos_data(self, request):
        form = SpotrebniKosForm(request.GET or None)

        rows = []
        circle_row = None
        maso_stat = None
        bio_stat = None
        volny_cukr_g = None
        legislativni_ukazatele = None
        kontroly_jidelnicku = None
        souhrn = None
        export_query = request.GET.urlencode()
        data_warnings = validace_surovin_pro_sk()

        if form.is_valid():
            date_from, date_to, label = form.get_period()
            stravovaci_skupina = form.cleaned_data.get("stravovaci_skupina")

            if stravovaci_skupina:
                qs = OrderItem.objects.filter(
                    order__datum_vydeje__gte=date_from,
                    order__datum_vydeje__lte=date_to,
                    order__user__stravovaci_skupina=stravovaci_skupina,
                )
                stravnikodny = qs.aggregate(celkem=Sum("quantity"))["celkem"] or 0

                rok = date_from.year
                mesic = date_from.month

                rows = priprav_radky_spotrebi_kos_tabulka(
                    rok=rok,
                    mesic=mesic,
                    stravovaci_skupina=stravovaci_skupina,
                    pocet_stravniku=int(stravnikodny),
                    date_from=date_from,
                    date_to=date_to,
                )

                for r in rows:
                    if r.get("norma_g") or r.get("skutecnost_g"):
                        circle_row = type(
                            "Row",
                            (),
                            {
                                "label": label,
                                "plneni": float(r["skutecnost_pct"]),
                            },
                        )()
                        break

                maso_stat = spocitej_podil_masnych_vyrobku(
                    date_from, date_to, stravovaci_skupina
                )
                bio_stat = spocitej_podil_bio(
                    date_from, date_to, stravovaci_skupina
                )
                volny_cukr_g = spocitej_volny_cukr(
                    date_from, date_to, stravovaci_skupina
                )
                legislativni_ukazatele = spocitej_legislativni_ukazatele_sk(
                    date_from, date_to, stravovaci_skupina
                )
                kontroly_jidelnicku = zkontroluj_jidelnicek_sk(
                    date_from, date_to, stravovaci_skupina
                )
                souhrn = spocitej_souhrn_spotrebniho_kose(
                    date_from, date_to, stravovaci_skupina
                )

        return {
            "title": "Spotřební koš – nové metodické ukazatele",
            "form": form,
            "rows": rows,
            "circle_row": circle_row,
            "maso_stat": maso_stat,
            "bio_stat": bio_stat,
            "volny_cukr_g": volny_cukr_g,
            "legislativni_ukazatele": legislativni_ukazatele,
            "kontroly_jidelnicku": kontroly_jidelnicku,
            "souhrn": souhrn,
            "export_query": export_query,
            "data_warnings": data_warnings,
        }

    def spotrebni_kos_xls_view(self, request):
        data = self._spotrebni_kos_data(request)
        souhrn = data.get("souhrn") or {}
        rows = data.get("rows") or []

        html = [
            "<html><head><meta charset='utf-8'></head><body>",
            "<h1>Spotřební koš</h1>",
            "<table border='1'>",
            "<tr><th>Ukazatel</th><th>Hodnota</th></tr>",
            f"<tr><td>Počet jídel</td><td>{souhrn.get('pocet_jidel', 0)}</td></tr>",
            f"<tr><td>Počet strávníků</td><td>{souhrn.get('pocet_stravniku', 0)}</td></tr>",
            f"<tr><td>Počet výdejek</td><td>{souhrn.get('pocet_vydejek', 0)}</td></tr>",
            "</table><br>",
            "<table border='1'>",
            "<tr><th>Skupina potravin</th><th>Norma [g]</th><th>Skutečnost [g]</th><th>Rozdíl [g]</th><th>Skutečnost [%]</th><th>Limit</th><th>Stav</th></tr>",
        ]
        for row in rows:
            max_pct = row.get("max_pct")
            limit = f"{row.get('min_pct', 0):.0f} %"
            limit += f" - {max_pct:.0f} %" if max_pct is not None else "+"
            html.append(
                "<tr>"
                f"<td>{row['skupina_nazev']}</td>"
                f"<td>{row['norma_g']:.2f}</td>"
                f"<td>{row['skutecnost_g']:.2f}</td>"
                f"<td>{row['rozdil_g']:.2f}</td>"
                f"<td>{row['skutecnost_pct']:.2f}</td>"
                f"<td>{limit}</td>"
                f"<td>{row['stav']}</td>"
                "</tr>"
            )
        html.append("</table></body></html>")

        response = HttpResponse("\n".join(html), content_type="application/vnd.ms-excel; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="spotrebni_kos.xls"'
        return response

    def spotrebni_kos_pdf_view(self, request):
        data = self._spotrebni_kos_data(request)
        souhrn = data.get("souhrn") or {}
        rows = data.get("rows") or []

        font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "DejaVuSans.ttf")
        pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))

        styles = getSampleStyleSheet()
        styles["Normal"].fontName = "DejaVuSans"
        styles["Title"].fontName = "DejaVuSans"
        styles["Heading2"].fontName = "DejaVuSans"

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
            leftMargin=1.2 * cm,
            rightMargin=1.2 * cm,
        )
        elements = [
            Paragraph("Spotřební koš", styles["Title"]),
            Spacer(1, 0.4 * cm),
            Paragraph(f"Počet jídel: {souhrn.get('pocet_jidel', 0)}", styles["Normal"]),
            Paragraph(f"Počet strávníků: {souhrn.get('pocet_stravniku', 0)}", styles["Normal"]),
            Paragraph(f"Počet výdejek: {souhrn.get('pocet_vydejek', 0)}", styles["Normal"]),
            Spacer(1, 0.5 * cm),
        ]

        table_data = [["Skupina", "Norma [g]", "Skutečnost [g]", "Rozdíl [g]", "%", "Limit", "Stav"]]
        for row in rows:
            max_pct = row.get("max_pct")
            limit = f"{row.get('min_pct', 0):.0f} %"
            limit += f" - {max_pct:.0f} %" if max_pct is not None else "+"
            table_data.append([
                row["skupina_nazev"],
                f"{row['norma_g']:.0f}",
                f"{row['skutecnost_g']:.0f}",
                f"{row['rozdil_g']:.0f}",
                f"{row['skutecnost_pct']:.2f}",
                limit,
                row["stav"],
            ])

        table = Table(table_data, colWidths=[4.0 * cm, 2.2 * cm, 2.4 * cm, 2.0 * cm, 1.6 * cm, 2.2 * cm, 2.3 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                    ("ALIGN", (0, 0), (0, -1), "LEFT"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
                ]
            )
        )
        elements.append(table)

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="spotrebni_kos.pdf"'
        return response


# -------------------------------------------------------------------
# Základní sklad – admin
# -------------------------------------------------------------------


@admin.register(Surovina)
class SurovinaAdmin(admin.ModelAdmin):
    list_display = (
        "nazev",
        "jednotka",
        "karta_suroviny_link",
        "skupina_sk",
        "koeficient_ciste_hmotnosti_sk",
        "koeficient_zapoctu_sk",
        "je_masny_vyrobek",
        "je_bio",
        "je_sezonni",
        "je_sterilovana_nebo_kompot",
        "hmotnost_ks_g_display",
        "prumerna_cena_za_jednotku",
    )
    list_filter = ("jednotka", "skupina_sk", "je_masny_vyrobek", "je_bio")
    search_fields = ("nazev",)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("<int:surovina_id>/karta/", self.admin_site.admin_view(self.karta_suroviny_view), name="sklad_surovina_karta"),
            path("<int:surovina_id>/karta/xlsx/", self.admin_site.admin_view(self.karta_suroviny_xlsx_view), name="sklad_surovina_karta_xlsx"),
            path("<int:surovina_id>/karta/pdf/", self.admin_site.admin_view(self.karta_suroviny_pdf_view), name="sklad_surovina_karta_pdf"),
        ]
        return custom_urls + urls

    def karta_suroviny_link(self, obj):
        return format_html('<a class="button" href="{}">Karta</a>', f"{obj.pk}/karta/")

    karta_suroviny_link.short_description = "Karta"

    def karta_suroviny_view(self, request, surovina_id):
        surovina = get_object_or_404(Surovina, pk=surovina_id)
        date_to = parse_date(request.GET.get("date_to") or "") or date.today()
        date_from = parse_date(request.GET.get("date_from") or "") or (date_to - timedelta(days=90))
        data = karta_suroviny_data(surovina, date_from=date_from, date_to=date_to)
        context = dict(
            self.admin_site.each_context(request),
            title=f"Karta suroviny: {surovina.nazev}",
            data=data,
        )
        return TemplateResponse(request, "admin/sklad/karta_suroviny.html", context)

    def _karta_suroviny_data_z_requestu(self, request, surovina_id):
        surovina = get_object_or_404(Surovina, pk=surovina_id)
        date_to = parse_date(request.GET.get("date_to") or "") or date.today()
        date_from = parse_date(request.GET.get("date_from") or "") or (date_to - timedelta(days=90))
        return karta_suroviny_data(surovina, date_from=date_from, date_to=date_to)

    def karta_suroviny_xlsx_view(self, request, surovina_id):
        data = self._karta_suroviny_data_z_requestu(request, surovina_id)
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Karta suroviny"
        ws.append([f"Karta suroviny: {data['surovina'].nazev}"])
        ws.append([f"Období: {data['date_from'].strftime('%d.%m.%Y')} - {data['date_to'].strftime('%d.%m.%Y')}"])
        ws.append([])
        ws.append(["Aktuální stav", data["stav_display"]])
        ws.append(["Minimum", data["minimum_display"]])
        ws.append(["Hodnota stavu Kč", float(data["hodnota_stavu"] or 0)])
        ws.append(["Spotřeba období", data["spotreba_obdobi_display"]])
        ws.append(["Náklady spotřeby Kč", float(data["naklady_spotreby"] or 0)])
        ws.append([])
        ws.append(["Datum", "Typ", "Množství", "Cena za jednotku", "Hodnota Kč", "Doklad", "Poznámka"])
        for pohyb in data["pohyby"]:
            doklad = "-"
            if pohyb.prijem_id:
                doklad = f"Příjemka #{pohyb.prijem_id}"
            elif pohyb.vydejka_id:
                doklad = f"Výdejka #{pohyb.vydejka_id}"
            elif pohyb.inventura_id:
                doklad = f"Inventura #{pohyb.inventura_id}"
            elif pohyb.odpis_expirace_id:
                doklad = f"Odpis expirace #{pohyb.odpis_expirace_id}"
            ws.append([
                pohyb.datum.strftime("%d.%m.%Y %H:%M"),
                pohyb.get_typ_display(),
                float(pohyb.mnozstvi or 0),
                float(pohyb.cena_za_jednotku or 0),
                float((pohyb.mnozstvi or Decimal("0")) * (pohyb.cena_za_jednotku or Decimal("0"))),
                doklad,
                pohyb.poznamka,
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="karta_suroviny_{data["surovina"].id}.xlsx"'
        return response

    def karta_suroviny_pdf_view(self, request, surovina_id):
        data = self._karta_suroviny_data_z_requestu(request, surovina_id)
        styles, font_name = _pdf_styles()
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.2 * cm, leftMargin=1.2 * cm, topMargin=1.4 * cm, bottomMargin=1.4 * cm)
        story = [
            Paragraph(f"Karta suroviny: {escape(data['surovina'].nazev)}", styles["Title"]),
            Spacer(1, 0.25 * cm),
            Paragraph(f"Období: {data['date_from'].strftime('%d.%m.%Y')} - {data['date_to'].strftime('%d.%m.%Y')}", styles["Normal"]),
            Paragraph(f"Aktuální stav: {data['stav_display']}", styles["Normal"]),
            Paragraph(f"Hodnota stavu: {data['hodnota_stavu']:.2f} Kč", styles["Normal"]),
            Paragraph(f"Spotřeba období: {data['spotreba_obdobi_display']} / {data['naklady_spotreby']:.2f} Kč", styles["Normal"]),
            Spacer(1, 0.35 * cm),
            Paragraph("Pohyby", styles["Heading2"]),
        ]
        table_data = [["Datum", "Typ", "Množství", "Cena", "Hodnota", "Doklad"]]
        for pohyb in data["pohyby"]:
            doklad = "-"
            if pohyb.prijem_id:
                doklad = f"Příjemka #{pohyb.prijem_id}"
            elif pohyb.vydejka_id:
                doklad = f"Výdejka #{pohyb.vydejka_id}"
            elif pohyb.inventura_id:
                doklad = f"Inventura #{pohyb.inventura_id}"
            elif pohyb.odpis_expirace_id:
                doklad = f"Odpis #{pohyb.odpis_expirace_id}"
            table_data.append([
                pohyb.datum.strftime("%d.%m.%Y"),
                pohyb.get_typ_display(),
                format_mnozstvi_s_jednotkou(data["surovina"], pohyb.mnozstvi),
                format_cena_za_jednotku(data["surovina"], pohyb.cena_za_jednotku or Decimal("0")),
                f"{((pohyb.mnozstvi or Decimal('0')) * (pohyb.cena_za_jednotku or Decimal('0'))):.2f} Kč",
                doklad,
            ])
        if len(table_data) == 1:
            table_data.append(["Bez pohybů", "-", "-", "-", "-", "-"])
        table = Table(table_data, colWidths=[2.4 * cm, 3.2 * cm, 3 * cm, 3.4 * cm, 2.5 * cm, 3.3 * cm], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)
        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="karta_suroviny_{data["surovina"].id}.pdf"'
        return response
    fieldsets = (
        (
            "Základní údaje",
            {
                "fields": ("nazev", "jednotka"),
            },
        ),
        (
            "Spotřební koš",
            {
                "fields": (
                    "skupina_sk",
                    "koeficient_ciste_hmotnosti_sk",
                    "koeficient_zapoctu_sk",
                    "koeficient_sk",
                    "je_masny_vyrobek",
                    "je_bio",
                    "je_sezonni",
                    "je_sterilovana_nebo_kompot",
                    "je_rostlinny_tuk",
                    "je_zivocisny_tuk",
                    "je_zakazano_pro_skolni_stravovani",
                    "podil_celozrnne_slozky",
                    "volny_cukr_na_100g",
                ),
            },
        ),
        (
            "Hmotnost a ceny",
            {
                "fields": (
                    "hmotnost_ks_g",
                    "prumerna_cena_za_jednotku",
                ),
            },
        ),
    )
    readonly_fields = ("prumerna_cena_za_jednotku",)

    def hmotnost_ks_g_display(self, obj):
        if obj.jednotka != "ks":
            return "-"
        if obj.hmotnost_ks_g is None:
            return "nenastaveno"
        return f"{obj.hmotnost_ks_g} g"

    hmotnost_ks_g_display.short_description = "g / ks"


@admin.register(StavSkladu)
class StavSkladuAdmin(admin.ModelAdmin):
    list_display = ("surovina", "mnozstvi_display", "min_mnozstvi_display", "skladova_jednotka_display")
    fields = ("surovina", "mnozstvi_display", "min_mnozstvi", "min_mnozstvi_display", "skladova_jednotka_display")
    readonly_fields = ("surovina", "mnozstvi_display", "min_mnozstvi_display", "skladova_jednotka_display")

    def mnozstvi_display(self, obj):
        return format_mnozstvi_s_jednotkou(obj.surovina, obj.mnozstvi)

    mnozstvi_display.short_description = "Množství"

    def min_mnozstvi_display(self, obj):
        return format_mnozstvi_s_jednotkou(obj.surovina, obj.min_mnozstvi)

    min_mnozstvi_display.short_description = "Minimální množství"

    def skladova_jednotka_display(self, obj):
        return obj.surovina.jednotka

    skladova_jednotka_display.short_description = "Skladová jednotka"

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(Dodavatel)
class DodavatelAdmin(admin.ModelAdmin):
    list_display = ("nazev", "ico", "dic", "email", "telefon", "aktivni")
    list_filter = ("aktivni",)
    search_fields = ("nazev", "ico", "dic", "email", "telefon")


@admin.register(PohybSkladu)
class PohybSkladuAdmin(admin.ModelAdmin):
    list_display = ("datum", "surovina", "typ", "mnozstvi_display", "sarze_skladu", "cena_za_jednotku_display", "hodnota_display", "doklad_link", "poznamka")
    list_filter = ("typ", "datum", "surovina")
    search_fields = ("surovina__nazev", "vydejka__id", "prijem__id", "poznamka")
    date_hierarchy = "datum"

    def mnozstvi_display(self, obj):
        return format_mnozstvi_s_jednotkou(obj.surovina, obj.mnozstvi)

    mnozstvi_display.short_description = "Množství"

    def cena_za_jednotku_display(self, obj):
        return format_cena_za_jednotku(obj.surovina, obj.cena_za_jednotku)

    cena_za_jednotku_display.short_description = "Cena za jednotku"

    def hodnota_display(self, obj):
        hodnota = (obj.mnozstvi or Decimal("0")) * (obj.cena_za_jednotku or Decimal("0"))
        return f"{hodnota:.2f} Kč"

    hodnota_display.short_description = "Hodnota"

    def doklad_link(self, obj):
        if obj.vydejka_id:
            url = f"/admin/sklad/vydejka/{obj.vydejka_id}/change/"
            return format_html('<a href="{}">Výdejka #{}</a>', url, obj.vydejka_id)
        if obj.prijem_id:
            url = f"/admin/sklad/prijemskladu/{obj.prijem_id}/change/"
            return format_html('<a href="{}">Příjem #{}</a>', url, obj.prijem_id)
        if getattr(obj, "inventura_id", None):
            url = f"/admin/sklad/inventura/{obj.inventura_id}/change/"
            return format_html('<a href="{}">Inventura #{}</a>', url, obj.inventura_id)
        if getattr(obj, "odpis_expirace_id", None):
            url = f"/admin/sklad/odpisexpirace/{obj.odpis_expirace_id}/change/"
            return format_html('<a href="{}">Odpis expirace #{}</a>', url, obj.odpis_expirace_id)
        return "-"

    doklad_link.short_description = "Doklad"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False


@admin.register(SarzeSkladu)
class SarzeSkladuAdmin(admin.ModelAdmin):
    list_display = (
        "surovina",
        "sarze",
        "typ_data_spotreby",
        "datum_spotreby",
        "mnozstvi_zbyva_display",
        "cena_za_jednotku_display",
        "stav",
    )
    list_filter = ("stav", "typ_data_spotreby", "datum_spotreby", "surovina")
    search_fields = ("surovina__nazev", "sarze", "poznamka")
    readonly_fields = (
        "surovina",
        "polozka_prijmu",
        "sarze",
        "typ_data_spotreby",
        "datum_spotreby",
        "mnozstvi_prijato",
        "mnozstvi_zbyva",
        "cena_za_jednotku",
        "stav",
        "poznamka",
    )
    date_hierarchy = "datum_spotreby"

    def mnozstvi_zbyva_display(self, obj):
        return format_mnozstvi_s_jednotkou(obj.surovina, obj.mnozstvi_zbyva)

    mnozstvi_zbyva_display.short_description = "Zbývá"

    def cena_za_jednotku_display(self, obj):
        return format_cena_za_jednotku(obj.surovina, obj.cena_za_jednotku)

    cena_za_jednotku_display.short_description = "Cena za jednotku"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.is_staff

    def has_delete_permission(self, request, obj=None):
        return False


class RecepturaPolozkaInline(admin.TabularInline):
    model = RecepturaPolozka
    extra = 1


class KomponentaSurovinaInline(admin.TabularInline):
    model = KomponentaSurovina
    extra = 1
    autocomplete_fields = ("surovina",)


class JidloKomponentaInline(admin.TabularInline):
    model = JidloKomponenta
    extra = 1
    autocomplete_fields = ("komponenta",)
    ordering = ("poradi", "id")


class PolozkaPrijmuInline(admin.StackedInline):
    model = PolozkaPrijmu
    extra = 1
    autocomplete_fields = ("surovina",)
    verbose_name = "Položka příjmu"
    verbose_name_plural = "Položky příjmu"
    fieldsets = (
        (
            "Surovina a balení",
            {
                "fields": (
                    "surovina",
                    ("pocet_baleni", "mnozstvi_v_baleni", "jednotka_baleni"),
                ),
                "description": (
                    "Zadej, co dodavatel přivezl. Např. 2 balení po 5 kg."
                ),
            },
        ),
        (
            "Cena",
            {
                "fields": (
                    ("cena_za_baleni_bez_dph", "sazba_dph"),
                ),
                "description": (
                    "Cena za jedno balení bez DPH. Jednotková skladová cena se dopočítá automaticky."
                ),
            },
        ),
        (
            "Šarže a trvanlivost",
            {
                "fields": (
                    ("sarze", "typ_data_spotreby", "datum_spotreby"),
                ),
            },
        ),
        (
            "Dopočtené hodnoty",
            {
                "classes": ("collapse",),
                "fields": (
                    ("mnozstvi", "jednotkova_cena"),
                    ("cena_za_baleni_s_dph", "cena_celkem_bez_dph", "cena_celkem_s_dph"),
                ),
                "description": (
                    "Tyto hodnoty se dopočítají po uložení položky."
                ),
            },
        ),
    )
    readonly_fields = (
        "cena_za_baleni_s_dph",
        "mnozstvi",
        "jednotkova_cena",
        "cena_celkem_bez_dph",
        "cena_celkem_s_dph",
    )

    def get_extra(self, request, obj=None, **kwargs):
        if obj and obj.uzavreny:
            return 0
        return super().get_extra(request, obj, **kwargs)

    def has_add_permission(self, request, obj=None):
        if obj and obj.uzavreny:
            return False
        return super().has_add_permission(request, obj)

    def has_change_permission(self, request, obj=None):
        if obj and (obj.uzavreny or obj.stornovano):
            return False
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.uzavreny:
            return False
        return super().has_delete_permission(request, obj)


class PolozkaInventuryInline(admin.TabularInline):
    model = PolozkaInventury
    extra = 0
    readonly_fields = ("stav_pred", "rozdil")
    autocomplete_fields = ("surovina",)

    def has_add_permission(self, request, obj=None):
        if obj and obj.uzavreny:
            return False
        return super().has_add_permission(request, obj)

    def has_change_permission(self, request, obj=None):
        if obj and (obj.uzavreny or obj.stornovano):
            return False
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.uzavreny:
            return False
        return super().has_delete_permission(request, obj)


class PohybSkladuInlineBase(admin.TabularInline):
    model = PohybSkladu
    extra = 0
    can_delete = False
    fields = ("datum", "surovina", "typ", "mnozstvi_display", "sarze_skladu", "cena_za_jednotku_display", "hodnota_display", "poznamka")
    readonly_fields = fields
    verbose_name = "Skladový pohyb"
    verbose_name_plural = "Skladové pohyby"

    def mnozstvi_display(self, obj):
        if not obj:
            return "-"
        return format_mnozstvi_s_jednotkou(obj.surovina, obj.mnozstvi)

    mnozstvi_display.short_description = "Množství"

    def cena_za_jednotku_display(self, obj):
        if not obj:
            return "-"
        return format_cena_za_jednotku(obj.surovina, obj.cena_za_jednotku)

    cena_za_jednotku_display.short_description = "Cena za jednotku"

    def hodnota_display(self, obj):
        if not obj:
            return "-"
        hodnota = (obj.mnozstvi or Decimal("0")) * (obj.cena_za_jednotku or Decimal("0"))
        return f"{hodnota:.2f} Kč"

    hodnota_display.short_description = "Hodnota"

    def has_add_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return True


class PohybPrijmuInline(PohybSkladuInlineBase):
    fk_name = "prijem"


class PohybVydejkyInline(PohybSkladuInlineBase):
    fk_name = "vydejka"


class PohybInventuryInline(PohybSkladuInlineBase):
    fk_name = "inventura"


class PohybOdpisuExpiraceInline(PohybSkladuInlineBase):
    fk_name = "odpis_expirace"


class PolozkaInventuryReadOnlyInline(admin.TabularInline):
    model = PolozkaInventury
    extra = 0
    can_delete = False
    readonly_fields = ("surovina", "stav_pred", "fyzicky_stav", "rozdil")
    fields = ("surovina", "stav_pred", "fyzicky_stav", "rozdil")

    def has_add_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False


@admin.register(KomponentaJidla)
class KomponentaJidlaAdmin(admin.ModelAdmin):
    list_display = ("nazev", "typ", "aktivni", "porce_text")
    list_filter = ("typ", "aktivni")
    search_fields = ("nazev",)
    inlines = [KomponentaSurovinaInline]


@admin.register(PrijemSkladu)
class PrijemSkladuAdmin(admin.ModelAdmin):
    list_display = ("id", "datum", "vytvoril", "uzavreny", "stornovano", "uzavren_at", "uzavrel")
    list_filter = ("uzavreny", "stornovano", "datum")
    inlines = [PolozkaPrijmuInline, PohybPrijmuInline]
    autocomplete_fields = ("dodavatel",)
    readonly_fields = (
        "stav_dokladu",
        "soucet_polozek_bez_dph_display",
        "soucet_polozek_s_dph_display",
        "rozdil_faktury_display",
        "validacni_varovani_display",
        "import_xlsx_link",
        "vytvoril",
        "uzavren_at",
        "uzavrel",
        "stornovano",
        "stornovano_at",
    )
    actions = ["stornovat_prijemky"]
    fieldsets = (
        (
            "Základní údaje",
            {
                "fields": ("datum", "dodavatel", "popis"),
            },
        ),
        (
            "Dodavatelský doklad",
            {
                "fields": (
                    "cislo_faktury",
                    "cislo_dodaciho_listu",
                    "datum_dodani",
                    "datum_vystaveni",
                    "datum_splatnosti",
                    "castka_faktury_celkem",
                    "priloha",
                ),
            },
        ),
        (
            "Součty",
            {
                "fields": (
                    "soucet_polozek_bez_dph_display",
                    "soucet_polozek_s_dph_display",
                    "rozdil_faktury_display",
                    "validacni_varovani_display",
                    "import_xlsx_link",
                ),
            },
        ),
        (
            "Stav a audit",
            {
                "fields": ("stav_dokladu", "uzavreny", "stornovano", "vytvoril", "uzavren_at", "uzavrel", "stornovano_at"),
            },
        ),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("<int:prijem_id>/import-xlsx/", self.admin_site.admin_view(self.import_xlsx_view), name="sklad_prijem_import_xlsx"),
            path("vzor-importu.xlsx", self.admin_site.admin_view(self.vzor_importu_xlsx_view), name="sklad_prijem_vzor_importu_xlsx"),
        ]
        return custom_urls + urls

    def stav_dokladu(self, obj):
        return _stav_dokladu_text(obj)

    stav_dokladu.short_description = "Stav dokladu"

    def soucet_polozek_bez_dph_display(self, obj):
        if not obj:
            return "-"
        return f"{obj.soucet_polozek_bez_dph:.2f} Kč"

    soucet_polozek_bez_dph_display.short_description = "Součet položek bez DPH"

    def soucet_polozek_s_dph_display(self, obj):
        if not obj:
            return "-"
        return f"{obj.soucet_polozek_s_dph:.2f} Kč"

    soucet_polozek_s_dph_display.short_description = "Součet položek s DPH"

    def rozdil_faktury_display(self, obj):
        if not obj or obj.rozdil_faktury is None:
            return "-"
        return f"{obj.rozdil_faktury:.2f} Kč"

    rozdil_faktury_display.short_description = "Rozdíl proti faktuře"

    def validacni_varovani_display(self, obj):
        if not obj or not obj.pk:
            return "Kontrola se zobrazí po uložení příjemky."
        varovani = validace_prijemky_pred_uzavrenim(obj)
        if not varovani:
            return format_html('<span class="text-success">Bez varování.</span>')
        return mark_safe("<ul>" + "".join(f"<li>{v}</li>" for v in varovani) + "</ul>")

    validacni_varovani_display.short_description = "Kontroly před uzavřením"

    def import_xlsx_link(self, obj):
        if not obj or not obj.pk:
            return "Import bude dostupný po uložení příjemky."
        if obj.uzavreny or obj.stornovano:
            return "Uzavřenou nebo stornovanou příjemku už nelze importem měnit."
        return format_html(
            '<a class="button" href="{}">Importovat položky z XLSX</a> '
            '<a class="button" href="{}">Stáhnout vzor</a>',
            reverse("admin:sklad_prijem_import_xlsx", args=[obj.pk]),
            reverse("admin:sklad_prijem_vzor_importu_xlsx"),
        )

    import_xlsx_link.short_description = "Import položek"

    def vzor_importu_xlsx_view(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Položky příjemky"
        ws.append([
            "surovina",
            "mnozstvi",
            "jednotkova_cena",
            "pocet_baleni",
            "mnozstvi_v_baleni",
            "jednotka_baleni",
            "cena_za_baleni_bez_dph",
            "sazba_dph",
            "sarze",
            "datum_spotreby",
            "typ_data_spotreby",
        ])
        ws.append(["Rýže dlouhozrnná", "10", "28.50", "2", "5", "kg", "142.50", "12", "RYZE-001", "2026-12-31", "MINIMALNI_TRVANLIVOST"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="vzor_importu_prijemky.xlsx"'
        return response

    def import_xlsx_view(self, request, prijem_id):
        prijem = get_object_or_404(PrijemSkladu, pk=prijem_id)
        if prijem.uzavreny or prijem.stornovano:
            messages.error(request, "Uzavřenou nebo stornovanou příjemku už nelze importem měnit.")
            return redirect(reverse("admin:sklad_prijemskladu_change", args=[prijem.pk]))

        form = ImportPrijemXlsxForm(request.POST or None, request.FILES or None)
        if request.method == "POST" and form.is_valid():
            from openpyxl import load_workbook

            wb = load_workbook(form.cleaned_data["soubor"], data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
            header_map = {name: index for index, name in enumerate(headers)}
            required = {"surovina", "mnozstvi", "jednotkova_cena"}
            if not required.issubset(header_map):
                messages.error(request, "Soubor nemá povinné sloupce: surovina, mnozstvi, jednotkova_cena.")
            else:
                vytvoreno = 0
                chyby = []
                for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    nazev = str(row[header_map["surovina"]] or "").strip()
                    if not nazev:
                        continue
                    surovina = Surovina.objects.filter(nazev__iexact=nazev).first()
                    if not surovina:
                        chyby.append(f"Řádek {row_index}: surovina '{nazev}' neexistuje.")
                        continue

                    def hodnota(sloupec, vychozi=None):
                        index = header_map.get(sloupec)
                        if index is None:
                            return vychozi
                        value = row[index]
                        return vychozi if value in (None, "") else value

                    datum_spotreby = hodnota("datum_spotreby")
                    if isinstance(datum_spotreby, datetime):
                        datum_spotreby = datum_spotreby.date()
                    elif isinstance(datum_spotreby, str):
                        datum_spotreby = parse_date(datum_spotreby)

                    PolozkaPrijmu.objects.create(
                        prijem=prijem,
                        surovina=surovina,
                        mnozstvi=Decimal(str(hodnota("mnozstvi", "0"))),
                        jednotkova_cena=Decimal(str(hodnota("jednotkova_cena", "0"))),
                        pocet_baleni=Decimal(str(hodnota("pocet_baleni", "1"))),
                        mnozstvi_v_baleni=Decimal(str(hodnota("mnozstvi_v_baleni"))) if hodnota("mnozstvi_v_baleni") not in (None, "") else None,
                        jednotka_baleni=str(hodnota("jednotka_baleni", "") or ""),
                        cena_za_baleni_bez_dph=Decimal(str(hodnota("cena_za_baleni_bez_dph"))) if hodnota("cena_za_baleni_bez_dph") not in (None, "") else None,
                        sazba_dph=Decimal(str(hodnota("sazba_dph", "0"))),
                        sarze=str(hodnota("sarze", "") or ""),
                        datum_spotreby=datum_spotreby,
                        typ_data_spotreby=str(hodnota("typ_data_spotreby", "POUZITELNOST") or "POUZITELNOST"),
                    )
                    vytvoreno += 1
                if vytvoreno:
                    messages.success(request, f"Naimportováno položek: {vytvoreno}.")
                for chyba in chyby[:10]:
                    messages.warning(request, chyba)
                return redirect(reverse("admin:sklad_prijemskladu_change", args=[prijem.pk]))

        context = dict(
            self.admin_site.each_context(request),
            title=f"Import položek příjemky #{prijem.id}",
            prijem=prijem,
            form=form,
            vzor_url=reverse("admin:sklad_prijem_vzor_importu_xlsx"),
        )
        return TemplateResponse(request, "admin/sklad/import_prijem_xlsx.html", context)

    def save_model(self, request, obj, form, change):
        if not obj.pk and not obj.vytvoril:
            obj.vytvoril = request.user
        super().save_model(request, obj, form, change)

    def save_form(self, request, form, change):
        obj = super().save_form(request, form, change)
        return _prepare_uzavreni_po_ulozeni(PrijemSkladu, obj, change)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance

        if getattr(obj, "_uzavrit_po_ulozeni", False):
            for varovani in validace_prijemky_pred_uzavrenim(obj):
                messages.warning(request, varovani)
            try:
                uzavri_prijem(obj, user=request.user)
            except ValidationError as exc:
                messages.error(request, " ".join(exc.messages))

    def get_readonly_fields(self, request, obj=None):
        ro = list(super().get_readonly_fields(request, obj))
        if obj and obj.uzavreny:
            ro += ["datum", "popis", "uzavreny", "stornovano"]
        return ro

    def has_delete_permission(self, request, obj=None):
        if obj and obj.uzavreny:
            return False
        return super().has_delete_permission(request, obj)

    def has_change_permission(self, request, obj=None):
        if obj and (obj.uzavreny or obj.stornovano):
            return False
        return super().has_change_permission(request, obj)

    @admin.action(description="Stornovat příjemky a vytvořit opačné pohyby")
    def stornovat_prijemky(self, request, queryset):
        stornovano = 0
        for prijem in queryset:
            try:
                if stornuj_prijem(prijem, user=request.user):
                    stornovano += 1
            except ValidationError as exc:
                messages.error(request, f"Příjemka #{prijem.id}: {' '.join(exc.messages)}")
        messages.success(request, f"Stornováno příjemek: {stornovano}.")


@admin.register(Inventura)
class InventuraAdmin(admin.ModelAdmin):
    list_display = ("id", "datum", "vytvoril", "uzavreny", "stornovano", "uzaverka_mesice_link", "inventurni_pdf_link", "uzavren_at", "uzavrel")
    list_filter = ("uzavreny", "stornovano", "datum")
    readonly_fields = (
        "stav_dokladu",
        "sarzova_inventura_link",
        "inventurni_pdf_link",
        "inventurni_nahled_display",
        "vytvoril",
        "uzavren_at",
        "uzavrel",
        "stornovano",
        "stornovano_at",
    )
    inlines = [PolozkaInventuryInline, PohybInventuryInline]
    actions = ["stornovat_inventury"]
    fieldsets = (
        (
            "Základní údaje",
            {
                "fields": ("datum", "popis"),
            },
        ),
        (
            "Stav a audit",
            {
                "fields": ("stav_dokladu", "sarzova_inventura_link", "inventurni_pdf_link", "inventurni_nahled_display", "uzavreny", "stornovano", "vytvoril", "uzavren_at", "uzavrel", "stornovano_at"),
            },
        ),
    )

    def stav_dokladu(self, obj):
        return _stav_dokladu_text(obj)

    stav_dokladu.short_description = "Stav dokladu"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("<int:inventura_id>/sarzova/", self.admin_site.admin_view(self.sarzova_inventura_view), name="sklad_inventura_sarzova"),
            path("<int:inventura_id>/pdf/", self.admin_site.admin_view(inventura_pdf_view), name="sklad_inventura_pdf"),
        ]
        return custom_urls + urls

    def sarzova_inventura_link(self, obj):
        if not obj or not obj.pk:
            return "Šaržová inventura bude dostupná po uložení dokladu."
        url = reverse("admin:sklad_inventura_sarzova", args=[obj.pk])
        return format_html('<a class="button" href="{}">Otevřít šaržovou inventuru</a>', url)

    sarzova_inventura_link.short_description = "Šaržové GUI"

    def inventurni_pdf_link(self, obj):
        if not obj or not obj.pk:
            return "PDF bude dostupné po uložení dokladu."
        url = reverse("admin:sklad_inventura_pdf", args=[obj.pk])
        return format_html('<a class="button" href="{}">Stáhnout PDF inventury</a>', url)

    inventurni_pdf_link.short_description = "PDF report"

    def uzaverka_mesice_link(self, obj):
        return _uzaverka_mesice_link(obj)

    uzaverka_mesice_link.short_description = "Uzávěrka měsíce"

    def sarzova_inventura_view(self, request, inventura_id):
        inventura = get_object_or_404(Inventura, pk=inventura_id)
        if request.method == "POST" and not inventura.uzavreny and not inventura.stornovano:
            action = request.POST.get("action")
            try:
                if action == "load":
                    count = napln_sarzovou_inventuru(inventura)
                    messages.success(request, f"Načteno šarží do inventury: {count}.")
                elif action == "save":
                    self._uloz_sarzovou_inventuru_z_postu(request, inventura)
                    messages.success(request, "Šaržová inventura byla uložena a přepočítána.")
                elif action == "close":
                    self._uloz_sarzovou_inventuru_z_postu(request, inventura)
                    uzavri_inventuru(inventura, user=request.user)
                    messages.success(request, "Inventura byla uzavřena podle šaržových položek.")
            except ValidationError as exc:
                messages.error(request, " ".join(exc.messages))

        souhrn = souhrn_sarzove_inventury(inventura) if inventura.sarze_polozky.exists() else {
            "radky": [],
            "pocet_polozek": 0,
            "manko": Decimal("0"),
            "prebytek": Decimal("0"),
            "cisty_rozdil": Decimal("0"),
        }
        context = dict(
            self.admin_site.each_context(request),
            title=f"Šaržová inventura #{inventura.id}",
            inventura=inventura,
            souhrn=souhrn,
            suroviny=Surovina.objects.order_by("nazev"),
            typy_data_spotreby=PolozkaPrijmu._meta.get_field("typ_data_spotreby").choices,
        )
        return TemplateResponse(request, "admin/sklad/inventura_sarzova.html", context)

    def _uloz_sarzovou_inventuru_z_postu(self, request, inventura):
        for row_id in request.POST.getlist("row_id"):
            pol = PolozkaInventurySarze.objects.get(pk=row_id, inventura=inventura)
            raw = request.POST.get(f"fyzicky_stav_{row_id}", "0") or "0"
            pol.fyzicky_stav = Decimal(raw.replace(",", "."))
            pol.poznamka = request.POST.get(f"poznamka_{row_id}", "")
            pol.save()

        for idx in range(1, 6):
            surovina_id = request.POST.get(f"new_surovina_{idx}")
            fyzicky_raw = request.POST.get(f"new_fyzicky_{idx}", "") or "0"
            fyzicky = Decimal(fyzicky_raw.replace(",", "."))
            if not surovina_id or fyzicky <= 0:
                continue
            surovina = Surovina.objects.get(pk=surovina_id)
            PolozkaInventurySarze.objects.create(
                inventura=inventura,
                surovina=surovina,
                sarze=request.POST.get(f"new_sarze_{idx}", ""),
                typ_data_spotreby=request.POST.get(f"new_typ_data_{idx}", "NEUVADI_SE") or "NEUVADI_SE",
                datum_spotreby=parse_date(request.POST.get(f"new_datum_{idx}", "") or ""),
                stav_pred=Decimal("0"),
                fyzicky_stav=fyzicky,
                cena_za_jednotku=Decimal((request.POST.get(f"new_cena_{idx}", "") or str(surovina.prumerna_cena_za_jednotku or 0)).replace(",", ".")),
                je_nova_sarze=True,
                poznamka=request.POST.get(f"new_poznamka_{idx}", ""),
            )
        synchronizuj_surovinove_polozky_inventury(inventura)

    def inventurni_nahled_display(self, obj):
        if not obj or not obj.pk:
            return "Náhled se zobrazí po uložení inventury."
        rows = inventurni_nahled(obj)
        if not rows:
            return "Inventura zatím nemá položky."
        html_rows = []
        for row in rows[:50]:
            sarze_text = ", ".join(
                f"{s.sarze or '#'+str(s.id)}: {format_mnozstvi_s_jednotkou(s.surovina, s.mnozstvi_zbyva)}"
                for s in row["sarze"][:5]
            ) or "bez šarží"
            html_rows.append(
                "<tr>"
                f"<td>{row['surovina'].nazev}</td>"
                f"<td>{row['stav_pred_display']}</td>"
                f"<td>{row['fyzicky_stav_display']}</td>"
                f"<td>{row['rozdil_display']}</td>"
                f"<td>{sarze_text}</td>"
                "</tr>"
            )
        return mark_safe(
            '<div class="table-responsive"><table class="table table-sm table-striped mb-0">'
            "<thead><tr><th>Surovina</th><th>Stav před</th><th>Fyzický stav</th><th>Rozdíl</th><th>Aktivní šarže</th></tr></thead>"
            f"<tbody>{''.join(html_rows)}</tbody></table></div>"
        )

    inventurni_nahled_display.short_description = "Šaržový náhled inventury"

    def has_delete_permission(self, request, obj=None):
        return False

    def save_model(self, request, obj, form, change):
        if not obj.pk and not obj.vytvoril:
            obj.vytvoril = request.user
        super().save_model(request, obj, form, change)

        if not change:
            self._napln_polozky_ze_stavu(obj)

    def save_form(self, request, form, change):
        obj = super().save_form(request, form, change)
        return _prepare_uzavreni_po_ulozeni(Inventura, obj, change)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance

        if getattr(obj, "_uzavrit_po_ulozeni", False):
            try:
                uzavri_inventuru(obj, user=request.user)
            except ValidationError as exc:
                messages.error(request, " ".join(exc.messages))

    def _napln_polozky_ze_stavu(self, inventura):
        suroviny = Surovina.objects.select_related("stav").all()
        polozky = []
        for s in suroviny:
            stav = getattr(s, "stav", None)
            stav_mnozstvi = stav.mnozstvi if stav else Decimal("0")
            polozky.append(
                PolozkaInventury(
                    inventura=inventura,
                    surovina=s,
                    stav_pred=stav_mnozstvi,
                    fyzicky_stav=stav_mnozstvi,
                    rozdil=Decimal("0"),
                )
            )
        PolozkaInventury.objects.bulk_create(polozky)

    def get_readonly_fields(self, request, obj=None):
        ro = list(super().get_readonly_fields(request, obj))
        if obj and obj.uzavreny:
            ro += [
                "datum",
                "dodavatel",
                "popis",
                "cislo_faktury",
                "cislo_dodaciho_listu",
                "datum_dodani",
                "datum_vystaveni",
                "datum_splatnosti",
                "castka_faktury_celkem",
                "priloha",
                "uzavreny",
                "stornovano",
            ]
        return ro

    def has_change_permission(self, request, obj=None):
        if obj and (obj.uzavreny or obj.stornovano):
            return False
        return super().has_change_permission(request, obj)

    @admin.action(description="Stornovat inventury a vytvořit opačné pohyby")
    def stornovat_inventury(self, request, queryset):
        stornovano = 0
        for inventura in queryset:
            try:
                if stornuj_inventuru(inventura, user=request.user):
                    stornovano += 1
            except ValidationError as exc:
                messages.error(request, f"Inventura #{inventura.id}: {' '.join(exc.messages)}")
        messages.success(request, f"Stornováno inventur: {stornovano}.")


@admin.register(OdpisExpirace)
class OdpisExpiraceAdmin(admin.ModelAdmin):
    list_display = ("id", "datum", "vytvoril", "uzavreny", "stornovano", "hodnota_celkem_display", "uzavren_at", "uzavrel")
    list_filter = ("uzavreny", "stornovano", "datum")
    readonly_fields = (
        "vytvoril",
        "uzavren_at",
        "uzavrel",
        "stornovano",
        "stornovano_at",
        "pocet_polozek_display",
        "mnozstvi_celkem_display",
        "hodnota_celkem_display",
    )
    inlines = [PohybOdpisuExpiraceInline]
    fieldsets = (
        ("Základní údaje", {"fields": ("datum", "popis")}),
        ("Vyhodnocení odpisu", {"fields": ("pocet_polozek_display", "mnozstvi_celkem_display", "hodnota_celkem_display")}),
        ("Stav a audit", {"fields": ("uzavreny", "stornovano", "vytvoril", "uzavren_at", "uzavrel", "stornovano_at")}),
    )

    def _souhrn(self, obj):
        if not obj or not obj.pk:
            return {"pocet_pohybu": 0, "mnozstvi_celkem": Decimal("0"), "hodnota_celkem": Decimal("0")}
        return souhrn_odpisu_expirace(obj)

    def pocet_polozek_display(self, obj):
        return self._souhrn(obj)["pocet_pohybu"]

    pocet_polozek_display.short_description = "Počet odepsaných položek"

    def mnozstvi_celkem_display(self, obj):
        return self._souhrn(obj)["mnozstvi_celkem"]

    mnozstvi_celkem_display.short_description = "Množství celkem"

    def hodnota_celkem_display(self, obj):
        return f"{self._souhrn(obj)['hodnota_celkem']:.2f} Kč"

    hodnota_celkem_display.short_description = "Hodnota odpisu"

    def save_model(self, request, obj, form, change):
        if not obj.pk and not obj.vytvoril:
            obj.vytvoril = request.user
        super().save_model(request, obj, form, change)

    def save_form(self, request, form, change):
        obj = super().save_form(request, form, change)
        return _prepare_uzavreni_po_ulozeni(OdpisExpirace, obj, change)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance
        if getattr(obj, "_uzavrit_po_ulozeni", False):
            try:
                uzavri_odpis_expirace(obj, user=request.user)
            except ValidationError as exc:
                messages.error(request, " ".join(exc.messages))

    def get_readonly_fields(self, request, obj=None):
        ro = list(super().get_readonly_fields(request, obj))
        if obj and obj.uzavreny:
            ro += ["datum", "popis", "uzavreny", "stornovano"]
        return ro

    def has_delete_permission(self, request, obj=None):
        if obj and obj.uzavreny:
            return False
        return super().has_delete_permission(request, obj)

    def has_change_permission(self, request, obj=None):
        if obj and (obj.uzavreny or obj.stornovano):
            return False
        return super().has_change_permission(request, obj)


@admin.register(SkladovaUzaverka)
class SkladovaUzaverkaAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "rok",
        "mesic",
        "uzavreny",
        "stornovano",
        "pripraveno_display",
        "konecny_stav_display",
        "rozdil_kontroly_display",
        "exporty_list_display",
        "uzavren_at",
        "uzavrel",
    )
    list_filter = ("rok", "mesic", "uzavreny", "stornovano")
    readonly_fields = (
        "stav_dokladu",
        "pocatecni_stav",
        "prijmy",
        "storna_prijmu",
        "vydeje",
        "storna_vydeju",
        "odpisy_expirace",
        "inventura_plus",
        "inventura_minus",
        "vypocet_konecneho_stavu",
        "konecny_stav",
        "rozdil_kontroly",
        "pruvodce_uzaverkou_display",
        "exporty_display",
        "vytvoril",
        "uzavren_at",
        "uzavrel",
        "stornovano",
        "stornovano_at",
    )
    actions = ["uzavrit_uzaverky", "stornovat_uzaverky"]
    fieldsets = (
        ("Období", {"fields": ("rok", "mesic", "datum", "popis")}),
        (
            "Hodnoty uzávěrky",
            {
                "fields": (
                    "pocatecni_stav",
                    "prijmy",
                    "storna_prijmu",
                    "vydeje",
                    "storna_vydeju",
                    "odpisy_expirace",
                    "inventura_plus",
                    "inventura_minus",
                    "vypocet_konecneho_stavu",
                    "konecny_stav",
                    "rozdil_kontroly",
                )
            },
        ),
        ("Průvodce uzávěrkou", {"fields": ("pruvodce_uzaverkou_display",)}),
        ("Exporty", {"fields": ("exporty_display",)}),
        ("Stav a audit", {"fields": ("stav_dokladu", "uzavreny", "stornovano", "vytvoril", "uzavren_at", "uzavrel", "stornovano_at")}),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("mesic/<int:rok>/<int:mesic>/", self.admin_site.admin_view(self.otevrit_mesicni_uzaverku), name="sklad_uzaverka_mesic"),
            path("<int:uzaverka_id>/xlsx/", self.admin_site.admin_view(self.export_xlsx), name="sklad_uzaverka_xlsx"),
            path("<int:uzaverka_id>/pdf/", self.admin_site.admin_view(self.export_pdf), name="sklad_uzaverka_pdf"),
        ]
        return custom_urls + urls

    def otevrit_mesicni_uzaverku(self, request, rok, mesic):
        uzaverka, created = SkladovaUzaverka.objects.get_or_create(
            rok=rok,
            mesic=mesic,
            defaults={
                "datum": date(rok, mesic, calendar.monthrange(rok, mesic)[1]),
                "vytvoril": request.user if request.user.is_authenticated else None,
                "popis": "Skladová uzávěrka připravená z inventurního dokladu.",
            },
        )
        if created:
            messages.success(request, f"Uzávěrka {mesic:02d}/{rok} byla založena.")
        else:
            messages.info(request, f"Uzávěrka {mesic:02d}/{rok} už existuje, otevírám ji.")
        return redirect(reverse("admin:sklad_skladovauzaverka_change", args=[uzaverka.pk]))

    def stav_dokladu(self, obj):
        return _stav_dokladu_text(obj)

    stav_dokladu.short_description = "Stav dokladu"

    def konecny_stav_display(self, obj):
        return f"{obj.konecny_stav:.2f} Kč"

    konecny_stav_display.short_description = "Konečná hodnota"

    def rozdil_kontroly_display(self, obj):
        return f"{obj.rozdil_kontroly:.2f} Kč"

    rozdil_kontroly_display.short_description = "Kontrolní rozdíl"

    def pripraveno_display(self, obj):
        if not obj or not obj.rok or not obj.mesic:
            return "-"
        data = pruvodce_skladovou_uzaverkou(obj.rok, obj.mesic)
        if data["pripraveno"]:
            return format_html('<span class="badge badge-success">Připraveno</span>')
        return format_html('<span class="badge badge-warning">Zkontrolovat</span>')

    pripraveno_display.short_description = "Připravenost"

    def exporty_list_display(self, obj):
        if not obj or not obj.pk:
            return "-"
        return format_html(
            '<a class="button" href="{}">PDF</a> <a class="button" href="{}">XLSX</a>',
            reverse("admin:sklad_uzaverka_pdf", args=[obj.pk]),
            reverse("admin:sklad_uzaverka_xlsx", args=[obj.pk]),
        )

    exporty_list_display.short_description = "Export"

    def exporty_display(self, obj):
        if not obj or not obj.pk:
            return "Export bude dostupný po uložení uzávěrky."
        return format_html(
            '<a class="button" href="{}">Stáhnout XLSX</a> '
            '<a class="button" href="{}">Stáhnout PDF</a>',
            f"{obj.pk}/xlsx/",
            f"{obj.pk}/pdf/",
        )

    exporty_display.short_description = "Exporty"

    def pruvodce_uzaverkou_display(self, obj):
        if not obj or not obj.rok or not obj.mesic:
            return "Průvodce se zobrazí po zadání roku a měsíce."
        data = pruvodce_skladovou_uzaverkou(obj.rok, obj.mesic)
        rows = []
        for kontrola in data["kontroly"]:
            badge = (
                '<span class="badge badge-success">V pořádku</span>'
                if kontrola["ok"] else
                '<span class="badge badge-danger">Vyžaduje kontrolu</span>'
            )
            rows.append(
                "<tr>"
                f"<td>{kontrola['nazev']}</td>"
                f"<td>{kontrola['pocet']}</td>"
                f"<td>{badge}</td>"
                "</tr>"
            )
        hlavicka = (
            '<div class="alert alert-success">Období je připravené k uzavření.</div>'
            if data["pripraveno"] else
            '<div class="alert alert-warning">Před uzavřením projdi označené kontroly.</div>'
        )
        return mark_safe(
            hlavicka
            + '<table class="table table-sm table-striped"><thead><tr><th>Kontrola</th><th>Počet</th><th>Stav</th></tr></thead>'
            + f"<tbody>{''.join(rows)}</tbody></table>"
            + f"<p><strong>Kontrolní rozdíl:</strong> {data['uzaverka']['rozdil_kontroly']:.2f} Kč</p>"
        )

    pruvodce_uzaverkou_display.short_description = "Průvodce uzávěrkou"

    def save_model(self, request, obj, form, change):
        if not obj.pk and not obj.vytvoril:
            obj.vytvoril = request.user
        if obj.rok and obj.mesic and not obj.datum:
            obj.datum = date(obj.rok, obj.mesic, calendar.monthrange(obj.rok, obj.mesic)[1])
        super().save_model(request, obj, form, change)

    def save_form(self, request, form, change):
        obj = super().save_form(request, form, change)
        return _prepare_uzavreni_po_ulozeni(SkladovaUzaverka, obj, change)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance
        if getattr(obj, "_uzavrit_po_ulozeni", False):
            try:
                uzavri_skladovou_uzaverku(obj, user=request.user)
            except ValidationError as exc:
                messages.error(request, " ".join(exc.messages))

    def uzavrit_uzaverky(self, request, queryset):
        pocet = 0
        for uzaverka in queryset:
            try:
                if uzavri_skladovou_uzaverku(uzaverka, user=request.user):
                    pocet += 1
            except ValidationError as exc:
                messages.error(request, f"{uzaverka}: {' '.join(exc.messages)}")
        self.message_user(request, f"Uzavřeno uzávěrek: {pocet}.", messages.SUCCESS)

    uzavrit_uzaverky.short_description = "Uzavřít vybrané skladové uzávěrky"

    def stornovat_uzaverky(self, request, queryset):
        pocet = 0
        for uzaverka in queryset:
            if otevri_skladovou_uzaverku(uzaverka, user=request.user):
                pocet += 1
        self.message_user(request, f"Stornováno uzávěrek: {pocet}.", messages.WARNING)

    stornovat_uzaverky.short_description = "Stornovat vybrané uzávěrky a otevřít období"

    def _export_rows(self, uzaverka):
        return [
            ("Počáteční hodnota skladu", uzaverka.pocatecni_stav),
            ("Příjmy", uzaverka.prijmy),
            ("Storna příjmů", uzaverka.storna_prijmu),
            ("Výdeje", uzaverka.vydeje),
            ("Storna výdejek", uzaverka.storna_vydeju),
            ("Odpisy expirací", uzaverka.odpisy_expirace),
            ("Inventurní přebytky", uzaverka.inventura_plus),
            ("Inventurní manka", uzaverka.inventura_minus),
            ("Vypočtená konečná hodnota", uzaverka.vypocet_konecneho_stavu),
            ("Skutečná konečná hodnota", uzaverka.konecny_stav),
            ("Kontrolní rozdíl", uzaverka.rozdil_kontroly),
        ]

    def export_xlsx(self, request, uzaverka_id):
        uzaverka = get_object_or_404(SkladovaUzaverka, pk=uzaverka_id)
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Skladová uzávěrka"
        ws.append([f"Skladová uzávěrka {uzaverka.mesic:02d}/{uzaverka.rok}"])
        ws.append([])
        ws.append(["Ukazatel", "Hodnota Kč"])
        for nazev, hodnota in self._export_rows(uzaverka):
            ws.append([nazev, float(hodnota or 0)])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="skladova_uzaverka_{uzaverka.rok}_{uzaverka.mesic:02d}.xlsx"'
        return response

    def export_pdf(self, request, uzaverka_id):
        uzaverka = get_object_or_404(SkladovaUzaverka, pk=uzaverka_id)
        pruvodce = pruvodce_skladovou_uzaverkou(uzaverka.rok, uzaverka.mesic)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
        styles, font_name = _pdf_styles()
        story = [
            Paragraph(f"Skladová uzávěrka {uzaverka.mesic:02d}/{uzaverka.rok}", styles["Title"]),
            Spacer(1, 0.4 * cm),
        ]
        data = [["Ukazatel", "Hodnota Kč"]]
        data += [[nazev, f"{hodnota:.2f} Kč"] for nazev, hodnota in self._export_rows(uzaverka)]
        table = Table(data, colWidths=[11 * cm, 5 * cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
        ]))
        story += [table, Spacer(1, 0.5 * cm), Paragraph("Kontrolní checklist", styles["Heading2"])]
        checklist = [["Kontrola", "Počet", "Stav"]]
        for kontrola in pruvodce["kontroly"]:
            checklist.append([
                kontrola["nazev"],
                kontrola["pocet"],
                "V pořádku" if kontrola["ok"] else "Vyžaduje kontrolu",
            ])
        checklist_table = Table(checklist, colWidths=[9 * cm, 2 * cm, 5 * cm], repeatRows=1)
        checklist_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
        ]))
        story.append(checklist_table)
        doc.build(story)
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="skladova_uzaverka_{uzaverka.rok}_{uzaverka.mesic:02d}.pdf"'
        return response


@admin.register(InventurniDoklad)
class InventurniDokladAdmin(admin.ModelAdmin):
    list_display = ("id", "datum", "vytvoril", "pocet_polozek", "uzaverka_mesice_link", "inventurni_pdf_link")
    list_filter = ("datum",)
    search_fields = ("id", "vytvoril__username")
    inlines = [PolozkaInventuryReadOnlyInline]
    readonly_fields = ("datum", "popis", "vytvoril", "uzavreny", "uzavren_at", "uzavrel", "inventurni_pdf_link")

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.filter(uzavreny=True)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return True

    def has_delete_permission(self, request, obj=None):
        return False

    def pocet_polozek(self, obj):
        return obj.polozky.count()

    pocet_polozek.short_description = "Počet položek"

    def inventurni_pdf_link(self, obj):
        if not obj or not obj.pk:
            return "-"
        url = reverse("admin:sklad_inventura_pdf", args=[obj.pk])
        return format_html('<a class="button" href="{}">Stáhnout PDF</a>', url)

    inventurni_pdf_link.short_description = "PDF report"

    def uzaverka_mesice_link(self, obj):
        return _uzaverka_mesice_link(obj)

    uzaverka_mesice_link.short_description = "Uzávěrka měsíce"


@admin.register(Vydejka)
class VydejkaAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "datum",
        "stravovaci_skupina",
        "typ_stravy",
        "uzavreny",
        "stornovano",
        "uzavren_at",
        "uzavrel",
    )
    list_filter = ("typ_stravy", "stravovaci_skupina", "datum", "uzavreny", "stornovano")
    search_fields = ("id", "stravovaci_skupina__nazev", "popis")
    readonly_fields = (
        "stav_dokladu",
        "nahled_cerpani_sarzi_display",
        "vytvoril",
        "uzavren_at",
        "uzavrel",
        "stornovano",
        "stornovano_at",
    )
    inlines = [PohybVydejkyInline]

    fieldsets = (
        (
            "Základní údaje",
            {
                "fields": ("datum", "stravovaci_skupina", "typ_stravy", "popis"),
            },
        ),
        (
            "Objednané receptury a suroviny",
            {
                "fields": ("nahled_cerpani_sarzi_display",),
                "description": "",
            },
        ),
        (
            "Stav a audit",
            {
                "fields": ("stav_dokladu", "uzavreny", "stornovano", "vytvoril", "uzavren_at", "uzavrel", "stornovano_at"),
            },
        ),
    )

    actions = [
        "akce_vygenerovat_z_objednavek",
        "uzavrit_vydejky",
        "stornovat_vydejky",
    ]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:vydejka_id>/pdf/",
                self.admin_site.admin_view(vydejka_pdf_view),
                name="sklad_vydejka_pdf",
            ),
        ]
        return custom_urls + urls

    def stav_dokladu(self, obj):
        return _stav_dokladu_text(obj)

    stav_dokladu.short_description = "Stav dokladu"

    def nahled_cerpani_sarzi_display(self, obj):
        if not obj or not obj.pk:
            return "Náhled se zobrazí po prvním uložení výdejky."
        data = nahled_vydejky(obj)
        if not data["radky"]:
            return "Výdejka zatím nemá položky."

        rows = []
        for row in data["radky"]:
            sarze = row.get("sarze")
            if row.get("chybi"):
                sarze_text = '<span class="badge badge-danger">Chybí použitelná šarže</span>'
                datum = "-"
            else:
                sarze_text = sarze.sarze or f"#{sarze.id}"
                datum = sarze.datum_spotreby.strftime("%d.%m.%Y") if sarze.datum_spotreby else "bez data"
            rows.append(
                "<tr>"
                f"<td>{row['surovina'].nazev}</td>"
                f"<td>{sarze_text}</td>"
                f"<td>{datum}</td>"
                f"<td>{format_mnozstvi_s_jednotkou(row['surovina'], row['mnozstvi'])}</td>"
                f"<td>{row['hodnota']:.2f} Kč</td>"
                "</tr>"
            )
        return mark_safe(
            '<div class="table-responsive"><table class="table table-sm table-striped mb-0">'
            "<thead><tr><th>Surovina</th><th>Šarže</th><th>Datum spotřeby</th><th>Množství</th><th>Hodnota</th></tr></thead>"
            f"<tbody>{''.join(rows)}</tbody>"
            f"<tfoot><tr><th colspan='4'>Odhadovaná hodnota výdeje</th><th>{data['hodnota_celkem']:.2f} Kč</th></tr></tfoot>"
            "</table></div>"
        )

    nahled_cerpani_sarzi_display.short_description = "Náhled čerpání šarží"

    def save_model(self, request, obj, form, change):
        if not obj.pk and not obj.vytvoril:
            obj.vytvoril = request.user
        super().save_model(request, obj, form, change)

    def save_form(self, request, form, change):
        obj = super().save_form(request, form, change)
        return _prepare_uzavreni_po_ulozeni(Vydejka, obj, change)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance

        if getattr(obj, "_uzavrit_po_ulozeni", False):
            _dopln_vydejku_z_objednavek_pokud_je_prazdna(obj)
            _upozorni_na_nedostatecne_stavy(request, obj)
            try:
                uzavri_vydejku(obj, user=request.user)
            except ValidationError as exc:
                messages.error(request, " ".join(exc.messages))

    @admin.action(description="Uzavřít výdejky a promítnout do skladu")
    def uzavrit_vydejky(self, request, queryset):
        uzavreno = 0
        for vydejka in queryset:
            if not vydejka.uzavreny:
                _dopln_vydejku_z_objednavek_pokud_je_prazdna(vydejka)
                _upozorni_na_nedostatecne_stavy(request, vydejka)
            try:
                if uzavri_vydejku(vydejka, user=request.user):
                    uzavreno += 1
            except ValidationError as exc:
                messages.error(request, f"Výdejka #{vydejka.id}: {' '.join(exc.messages)}")

        self.message_user(
            request,
            f"Uzavřeno a promítnuto do skladu: {uzavreno} výdejek.",
        )

    @admin.action(description="Vygenerovat / přepočítat z objednávek pro zvolené výdejky")
    def akce_vygenerovat_z_objednavek(self, request, queryset):
        pocet = 0
        for vydejka in queryset:
            if vydejka.uzavreny:
                continue
            generate_vydejka_from_orders(
                datum=vydejka.datum,
                stravovaci_skupina=vydejka.stravovaci_skupina,
                typ_stravy=vydejka.typ_stravy,
            )
            pocet += 1

        messages.success(
            request,
            f"Výdejky byly vygenerovány/přepočítány z objednávek ({pocet} ks).",
        )

    @admin.action(description="Stornovat výdejky a vrátit suroviny na sklad")
    def stornovat_vydejky(self, request, queryset):
        stornovano = 0
        for vydejka in queryset:
            try:
                if stornuj_vydejku(vydejka, user=request.user):
                    stornovano += 1
            except ValidationError as exc:
                messages.error(request, f"Výdejka #{vydejka.id}: {' '.join(exc.messages)}")
        messages.success(request, f"Stornováno výdejek: {stornovano}.")

    def get_readonly_fields(self, request, obj=None):
        ro = list(super().get_readonly_fields(request, obj))
        if obj and obj.uzavreny:
            ro += ["datum", "stravovaci_skupina", "typ_stravy", "popis", "uzavreny", "stornovano"]
        return ro

    def has_delete_permission(self, request, obj=None):
        if obj and (obj.uzavreny or obj.stornovano):
            return False
        return super().has_delete_permission(request, obj)

    def has_change_permission(self, request, obj=None):
        if obj and (obj.uzavreny or obj.stornovano):
            return False
        return super().has_change_permission(request, obj)

    def get_fieldsets(self, request, obj=None):
        fieldsets = list(super().get_fieldsets(request, obj))

        if obj and obj.pk:
            title, opts = fieldsets[1]
            opts = dict(opts)

            pdf_button = f"""
                <a href="/admin/sklad/vydejka/{obj.id}/pdf/"
                   class="btn btn-primary"
                   style="margin-bottom: 1em; display: inline-block;">
                    📄 Stáhnout PDF výdejky
                </a>
            """

            opts["description"] = mark_safe(pdf_button + self.objednavky_rekap(obj))
            fieldsets[1] = (title, opts)

        return fieldsets

    def objednavky_rekap(self, obj):
        from django.utils.html import escape

        if not obj or not obj.pk:
            return "Ulož výdejku, aby bylo co spočítat."

        order_items = get_order_items_for_vydejka(obj)

        debug = [
            f"datum: {obj.datum}",
            f"typ_stravy: {obj.typ_stravy}",
            f"stravovaci_skupina: {obj.stravovaci_skupina}",
            f"order_items po filtrech: {order_items.count()}",
        ]

        if not order_items.exists():
            return mark_safe(
                "<br>".join(escape(line) for line in debug)
                + "<br><strong>Pro tento filtr nejsou žádné objednávky.</strong>"
            )

        porce_per_jidlo, jidla, detail = objednavky_rekap_data(obj)

        bloky = []

        for jidlo_id, pocet_porci in porce_per_jidlo.items():
            jidlo = jidla[jidlo_id]

            header_html = (
                "<h4 style='margin-top: 1em;'>"
                f"{jidlo.nazev} "
                "<br/>"
                f"<span style='font-weight: normal; color: #666;'>(porcí: {pocet_porci})</span>"
                "</h4>"
            )

            komponenty_html = []

            for blok in detail.get(jidlo_id, []):
                komponenta = blok["komponenta"]
                radky = blok["radky"]

                if komponenta:
                    komponenty_html.append(
                        f"<h5 style='margin: .5em 0; color:#444;'>{escape(komponenta.nazev)}</h5>"
                    )

                if radky:
                    rows_html = []
                    for row in radky:
                        rows_html.append(
                            "<tr>"
                            f"<td>{escape(row['surovina'].nazev)}</td>"
                            f"<td style='text-align:right;'>{row['na_porci']} {escape(row['surovina'].jednotka)}</td>"
                            f"<td style='text-align:right;'>{row['celkem']} {escape(row['surovina'].jednotka)}</td>"
                            "</tr>"
                        )

                    komponenty_html.append(
                        '<div style="width:100%; max-width:none; overflow-x:auto;">'
                        '<table class="table" style="width:100%; max-width:none; table-layout:auto; margin-bottom: 1em;">'
                        "<thead>"
                        "<tr>"
                        "<th style='width:40%;'>Surovina</th>"
                        '<th style="text-align:right; width:30%;">Na 1 porci</th>'
                        '<th style="text-align:right; width:30%;">Celkem pro všechny porce</th>'
                        "</tr>"
                        "</thead>"
                        "<tbody>"
                        + "".join(rows_html)
                        + "</tbody>"
                        "</table>"
                        "</div>"
                    )
                else:
                    komponenty_html.append("<p style='color:#888;'>Komponenta nemá suroviny.</p>")

            bloky.append(header_html + "".join(komponenty_html))

        return mark_safe(
            "<div style='width:100%; max-width:none;'>"
            + "".join(bloky)
            + "</div>"
        )


def vydejka_pdf_view(request, vydejka_id):
    from django.utils.html import escape

    vydejka = get_object_or_404(Vydejka, pk=vydejka_id)

    font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "DejaVuSans.ttf")
    pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))

    styles = getSampleStyleSheet()
    styles["Normal"].fontName = "DejaVuSans"
    styles["Title"].fontName = "DejaVuSans"
    styles["Heading2"].fontName = "DejaVuSans"

    order_items = get_order_items_for_vydejka(vydejka)

    if not order_items.exists():
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm
        )
        elements = []

        sk = (
            vydejka.stravovaci_skupina.kod
            if vydejka.stravovaci_skupina
            else "bez skupiny"
        )
        elements.append(
            Paragraph(
                f"Výdejka #{vydejka.id} – {vydejka.datum.strftime('%d.%m.%Y')} ({sk})",
                styles["Title"],
            )
        )
        elements.append(Spacer(1, 0.5 * cm))
        elements.append(
            Paragraph("Pro tento filtr nejsou žádné objednávky.", styles["Normal"])
        )

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="vydejka_{vydejka.id}.pdf"'
        return response

    porce_per_jidlo, jidla, detail = objednavky_rekap_data(vydejka)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
    )
    elements = []

    sk = (
        vydejka.stravovaci_skupina.kod
        if vydejka.stravovaci_skupina
        else "bez skupiny"
    )
    nadpis = f"Výdejka #{vydejka.id} – {vydejka.datum.strftime('%d.%m.%Y')} ({sk})"
    elements.append(Paragraph(nadpis, styles["Title"]))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(
        Paragraph(
            f"Typ stravy: {vydejka.get_typ_stravy_display()}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 0.7 * cm))

    for jidlo_id, pocet_porci in porce_per_jidlo.items():
        jidlo = jidla[jidlo_id]

        elements.append(
            Paragraph(
                f"<b>{escape(jidlo.nazev)}</b> – {pocet_porci} ks",
                styles["Heading2"],
            )
        )
        elements.append(Spacer(1, 0.2 * cm))

        for blok in detail.get(jidlo_id, []):
            komponenta = blok["komponenta"]
            radky = blok["radky"]

            if komponenta:
                elements.append(
                    Paragraph(
                        f"<b>{escape(komponenta.nazev)}</b>",
                        styles["Normal"],
                    )
                )
                elements.append(Spacer(1, 0.15 * cm))

            if radky:
                data = [["Surovina", "Na 1 porci", "Celkem"]]
                for row in radky:
                    surovina = row["surovina"]
                    data.append([
                        surovina.nazev,
                        f"{row['na_porci']:.3f} {surovina.jednotka}",
                        f"{row['celkem']:.3f} {surovina.jednotka}",
                    ])

                table = Table(data, colWidths=[7 * cm, 4 * cm, 5 * cm])
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                            ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans"),
                            ("FONTNAME", (0, 1), (-1, -1), "DejaVuSans"),
                            ("FONTSIZE", (0, 0), (-1, -1), 9),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ]
                    )
                )
                elements.append(table)
            else:
                elements.append(
                    Paragraph(
                        "<i>Komponenta nemá vyplněné suroviny.</i>",
                        styles["Normal"],
                    )
                )

            elements.append(Spacer(1, 0.4 * cm))

        elements.append(Spacer(1, 0.3 * cm))

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="vydejka_{vydejka.id}.pdf"'
    return response
